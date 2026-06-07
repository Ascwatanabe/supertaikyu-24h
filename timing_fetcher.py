#!/usr/bin/env python3
"""Super Taikyu live timing fetcher."""

from __future__ import annotations

import argparse
import csv
import functools
import html
import http.server
import json
import sys
import threading
import time
import webbrowser
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

APP_DIR = Path(__file__).resolve().parent
BASE_URL = "https://www.supertaikyu.live/json"
DEFAULT_CLASS = "ST-5F"
DEFAULT_INTERVAL = 5
DEFAULT_OUTPUT_DIR = APP_DIR / "data"
DEFAULT_HTTP_PORT = 8765
REQUEST_TIMEOUT = 10
MAX_RETRIES = 3

TEAM_NAME = "AndLegal Racing"
TEAM_TAGLINE = "ONE LAP AHEAD — ともに先へ"
TEAM_X_URL = "https://x.com/AndLegal_Racing"
TEAM_YOUTUBE_VIDEO_ID = "RDINupchH9Y"
TEAM_CAR_NOS = ("821", "822")
TEAM_LOGO_ASSET = "../assets/team-logo.jpg"
TEAM_BANNER_ASSET = "../assets/team-banner.jpg"

ALL_CLASSES = [
    "ST-X",
    "ST-Z",
    "ST-TCR",
    "ST-USA",
    "ST-Q",
    "ST-1",
    "ST-2",
    "ST-3",
    "ST-4",
    "ST-5F",
    "ST-5R",
]


@dataclass
class TimingRow:
    pos: str
    pic: str
    car_no: str
    car_class: str
    laps: str
    current_driver: str
    all_drivers: str
    team_name: str
    car_name: str
    best_lap: str
    last_lap: str
    s1: str
    s2: str
    s3: str
    best_lap_ms: int
    last_lap_ms: int
    s1_ms: int
    s2_ms: int
    s3_ms: int
    has_live_data: bool


@dataclass
class DriverRow:
    driver_name: str
    driver_slot: str
    car_no: str
    car_class: str
    team_name: str
    car_name: str
    best_lap: str
    last_lap: str
    s1: str
    s2: str
    s3: str
    best_lap_ms: int
    last_lap_ms: int
    s1_ms: int
    s2_ms: int
    s3_ms: int
    is_current: bool
    pos: str
    pic: str
    laps: str


@dataclass
class LapHistoryEntry:
    car_no: str
    car_class: str
    driver_slot: str
    driver_name: str
    lap_no: int
    lap_time_ms: int
    lap_time: str
    recorded_at: str


def class_label_for(class_filter: str | None) -> str:
    return class_filter or "ALL"


def filter_rows_by_class(rows: list[TimingRow], class_filter: str | None) -> list[TimingRow]:
    if not class_filter:
        return rows
    return [row for row in rows if row.car_class == class_filter]


def collect_available_classes(
    rows: list[TimingRow],
    driver_rows: list[DriverRow],
    lap_history: list[dict[str, Any]],
) -> list[str]:
    classes: set[str] = set()
    for row in rows:
        if row.car_class and row.car_class != "-":
            classes.add(row.car_class)
    for row in driver_rows:
        if row.car_class and row.car_class != "-":
            classes.add(row.car_class)
    for lap in lap_history:
        car_class = str(lap.get("car_class", "")).strip()
        if car_class and car_class != "-":
            classes.add(car_class)

    ordered = [cls for cls in ALL_CLASSES if cls in classes]
    for cls in sorted(classes):
        if cls not in ordered:
            ordered.append(cls)
    return ordered


def read_json_file(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}


def write_json_file(path: Path, payload: dict[str, Any]) -> None:
    ensure_output_dir(path.parent)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def latest_json_path(output_dir: Path) -> Path:
    return output_dir / "latest.json"


def latest_csv_path(output_dir: Path) -> Path:
    return output_dir / "latest.csv"


def history_csv_path(output_dir: Path) -> Path:
    return output_dir / "history.csv"


def lap_history_path(output_dir: Path) -> Path:
    return output_dir / "lap_history.json"


def lap_state_path(output_dir: Path) -> Path:
    return output_dir / "lap_state.json"


def sector_state_path(output_dir: Path) -> Path:
    return output_dir / "sector_state.json"


def view_data_path(output_dir: Path) -> Path:
    return output_dir / "view_data.json"


def warehouse_dir(output_dir: Path) -> Path:
    return output_dir / "warehouse"


def raw_snapshots_csv_path(output_dir: Path) -> Path:
    return warehouse_dir(output_dir) / "raw_snapshots.csv"


def drivers_running_csv_path(output_dir: Path) -> Path:
    return warehouse_dir(output_dir) / "drivers_running.csv"


def live_xlsx_path(output_dir: Path) -> Path:
    return output_dir / "timing_live.xlsx"


def make_lap_entry(
    car_no: str,
    car_class: str,
    driver_slot: str,
    driver_name: str,
    lap_no: int,
    lap_time_ms: int,
    recorded_at: str | None = None,
) -> dict[str, Any]:
    return asdict(
        LapHistoryEntry(
            car_no=car_no,
            car_class=car_class,
            driver_slot=driver_slot,
            driver_name=driver_name,
            lap_no=lap_no,
            lap_time_ms=lap_time_ms,
            lap_time=ms_to_laptime(lap_time_ms),
            recorded_at=recorded_at or datetime.now().isoformat(timespec="seconds"),
        )
    )


def bootstrap_lap_history_for_car(
    history_laps: list[dict[str, Any]],
    existing_keys: set[tuple[str, int]],
    car_no: str,
    car_class: str,
    laps: int,
    driver_slot: str,
    driver_name: str,
    lap_history_10: list[Any],
) -> None:
    if laps <= 0 or not lap_history_10:
        return

    valid_times = [int(value or 0) for value in lap_history_10 if int(value or 0) > 0]
    if not valid_times:
        return

    start_lap = max(1, laps - len(lap_history_10) + 1)
    for index, lap_time_ms in enumerate(lap_history_10):
        lap_time_ms = int(lap_time_ms or 0)
        if lap_time_ms <= 0:
            continue

        lap_no = start_lap + index
        if lap_no <= 0 or lap_no > laps:
            continue

        key = (car_no, lap_no)
        if key in existing_keys:
            continue

        history_laps.append(
            make_lap_entry(
                car_no, car_class, driver_slot, driver_name, lap_no, lap_time_ms
            )
        )
        existing_keys.add(key)


def update_lap_history(
    output_dir: Path,
    master: dict[str, Any],
    live: dict[str, Any],
    use_english: bool = False,
) -> Path:
    entry_info: dict[str, Any] = master.get("EntryInfo", {})
    lap_history_10_data: dict[str, Any] = live.get("LapHistory10Data", {})
    history_path = lap_history_path(output_dir)
    state_path = lap_state_path(output_dir)

    history_doc = read_json_file(history_path)
    history_laps: list[dict[str, Any]] = list(history_doc.get("laps", []))
    state: dict[str, Any] = dict(read_json_file(state_path))
    existing_keys = {(lap["car_no"], int(lap["lap_no"])) for lap in history_laps}

    for live_row in live.get("LiveData", []):
        car_no = str(live_row.get("CarNo", "")).strip()
        entry = entry_info.get(car_no)
        if not entry:
            continue

        car_class = str(entry.get("ClassStr", "")).strip() or "-"
        laps = int(live_row.get("LAPS", 0) or 0)
        lap_time_ms = int(live_row.get("LapTime", 0) or 0)
        driver_slot = str(live_row.get("Driver", "")).strip()
        driver_name = get_driver_name(entry, driver_slot, use_english) or driver_slot or "-"
        prev = state.get(car_no)

        if prev is None:
            bootstrap_lap_history_for_car(
                history_laps,
                existing_keys,
                car_no,
                car_class,
                laps,
                driver_slot,
                driver_name,
                list(lap_history_10_data.get(car_no, [])),
            )

        if laps > 0 and lap_time_ms > 0:
            key = (car_no, laps)
            if key not in existing_keys:
                history_laps.append(
                    make_lap_entry(
                        car_no,
                        car_class,
                        driver_slot,
                        driver_name,
                        laps,
                        lap_time_ms,
                    )
                )
                existing_keys.add(key)
            else:
                for lap in history_laps:
                    if lap["car_no"] == car_no and int(lap["lap_no"]) == laps:
                        if int(lap["lap_time_ms"]) != lap_time_ms:
                            lap["lap_time_ms"] = lap_time_ms
                            lap["lap_time"] = ms_to_laptime(lap_time_ms)
                            lap["recorded_at"] = datetime.now().isoformat(timespec="seconds")
                        if not lap.get("car_class") or lap["car_class"] == "-":
                            lap["car_class"] = car_class
                        break

        state[car_no] = {
            "laps": laps,
            "lap_time_ms": lap_time_ms,
            "driver_slot": driver_slot,
            "car_class": car_class,
        }

    history_laps.sort(key=lambda lap: (lap.get("car_class", ""), lap["car_no"], int(lap["lap_no"])))
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "laps": history_laps,
    }
    write_json_file(history_path, payload)
    write_json_file(state_path, state)
    return history_path


def save_view_data_json(
    output_dir: Path,
    rows: list[TimingRow],
    driver_rows: list[DriverRow],
    master: dict[str, Any],
    live: dict[str, Any],
    interval: int,
    default_class: str,
) -> Path:
    history_doc = read_json_file(lap_history_path(output_dir))
    lap_history = history_doc.get("laps", [])
    payload = {
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "update_time_display": format_update_time(live.get("UpdateTime", "")),
        "status": race_status_text(live),
        "race_name": master.get("RaceNameL", ""),
        "race_year": master.get("RaceYear", ""),
        "default_class": default_class,
        "classes": collect_available_classes(rows, driver_rows, lap_history),
        "interval": interval,
        "cars": [asdict(row) for row in rows],
        "drivers": [asdict(row) for row in driver_rows],
        "lap_history": lap_history,
    }
    path = view_data_path(output_dir)
    write_json_file(path, payload)
    return path


def ms_to_laptime(ms: int | float | str) -> str:
    """Convert milliseconds to the same format used on supertaikyu.live."""
    try:
        time_ms = int(ms)
    except (TypeError, ValueError):
        return "-"

    if time_ms <= 0:
        return "-"

    hh = time_ms // 3_600_000
    mm = (time_ms % 3_600_000) // 60_000
    ss = (time_ms % 60_000) // 1_000
    milli = time_ms % 1_000

    if hh > 0:
        return f"{hh}:{mm:02d}'{ss:02d}.{milli:03d}"
    if mm > 0:
        return f"{mm}'{ss:02d}.{milli:03d}"
    return f"{ss}.{milli:03d}"


def _sector_value(values: list[Any], index: int) -> int:
    if index >= len(values):
        return 0
    try:
        return int(values[index] or 0)
    except (TypeError, ValueError):
        return 0


def complete_sector_times_ms(live_row: dict[str, Any]) -> tuple[int, int, int] | None:
    s2s = live_row.get("Sect2SectTime") or []
    s1_ms = _sector_value(s2s, 0)
    s2_ms = _sector_value(s2s, 1)
    s3_ms = _sector_value(s2s, 2)
    lap_time_ms = int(live_row.get("LapTime", 0) or 0)

    if s1_ms <= 0 or s2_ms <= 0 or s3_ms <= 0 or lap_time_ms <= 0:
        return None
    if abs((s1_ms + s2_ms + s3_ms) - lap_time_ms) > 10:
        return None
    return s1_ms, s2_ms, s3_ms


def update_sector_state(output_dir: Path, live: dict[str, Any]) -> Path:
    path = sector_state_path(output_dir)
    state: dict[str, Any] = dict(read_json_file(path))

    for live_row in live.get("LiveData", []):
        car_no = str(live_row.get("CarNo", "")).strip()
        if not car_no:
            continue
        complete = complete_sector_times_ms(live_row)
        if not complete:
            continue
        s1_ms, s2_ms, s3_ms = complete
        state[car_no] = {
            "s1_ms": s1_ms,
            "s2_ms": s2_ms,
            "s3_ms": s3_ms,
            "lap_time_ms": int(live_row.get("LapTime", 0) or 0),
        }

    write_json_file(path, state)
    return path


def sector_times_ms_from_live(
    live_row: dict[str, Any],
    sector_cache: dict[str, Any] | None = None,
) -> tuple[int, int, int]:
    complete = complete_sector_times_ms(live_row)
    if complete:
        return complete

    lap_time_ms = int(live_row.get("LapTime", 0) or 0)
    if sector_cache and int(sector_cache.get("lap_time_ms", 0) or 0) == lap_time_ms:
        return (
            int(sector_cache.get("s1_ms", 0) or 0),
            int(sector_cache.get("s2_ms", 0) or 0),
            int(sector_cache.get("s3_ms", 0) or 0),
        )

    return 0, 0, 0


def sector_times_from_live(
    live_row: dict[str, Any],
    sector_cache: dict[str, Any] | None = None,
) -> tuple[str, str, str, int, int, int]:
    s1_ms, s2_ms, s3_ms = sector_times_ms_from_live(live_row, sector_cache)
    return (
        ms_to_laptime(s1_ms),
        ms_to_laptime(s2_ms),
        ms_to_laptime(s3_ms),
        s1_ms,
        s2_ms,
        s3_ms,
    )


def fetch_json(session: requests.Session, name: str) -> dict[str, Any]:
    url = f"{BASE_URL}/{name}"
    last_error: Exception | None = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = session.get(url, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt < MAX_RETRIES:
                time.sleep(1)

    raise RuntimeError(f"Failed to fetch {url}: {last_error}") from last_error


def get_driver_name(
    entry: dict[str, Any] | None,
    driver_id: str,
    use_english: bool = False,
) -> str:
    if not entry:
        return ""

    drivers = entry.get("Drivers", {})
    driver = drivers.get(driver_id, {})
    key = "eNameL" if use_english else "jNameL"
    return str(driver.get(key, "")).strip()


def get_all_driver_names(entry: dict[str, Any] | None, use_english: bool = False) -> str:
    if not entry:
        return ""

    drivers = entry.get("Drivers", {})
    key = "eNameL" if use_english else "jNameL"
    names: list[str] = []

    for slot in sorted(drivers.keys()):
        name = str(drivers[slot].get(key, "")).strip()
        if name:
            names.append(name)

    return " / ".join(names)


def has_active_timing(live_data: list[dict[str, Any]]) -> bool:
    for row in live_data:
        if int(row.get("BestLapTime", 0) or 0) > 0:
            return True
        if int(row.get("LapTime", 0) or 0) > 0:
            return True
        if str(row.get("LAPS", "")).strip().isdigit() and int(row.get("LAPS", 0)) > 0:
            return True
    return False


def build_rows(
    master: dict[str, Any],
    live: dict[str, Any],
    class_filter: str | None,
    use_english: bool = False,
    sector_state: dict[str, Any] | None = None,
) -> list[TimingRow]:
    entry_info: dict[str, Any] = master.get("EntryInfo", {})
    live_data: list[dict[str, Any]] = live.get("LiveData", [])
    live_by_car = {str(row.get("CarNo", "")): row for row in live_data}

    rows: list[TimingRow] = []

    for car_no, entry in entry_info.items():
        car_class = str(entry.get("ClassStr", "")).strip()
        if class_filter and car_class != class_filter:
            continue

        live_row = live_by_car.get(str(car_no), {})
        current_driver_id = str(live_row.get("Driver", "")).strip()
        current_driver = get_driver_name(entry, current_driver_id, use_english)

        best_lap_ms = int(live_row.get("BestLapTime", 0) or 0)
        last_lap_ms = int(live_row.get("LapTime", 0) or 0)
        sector_cache = (sector_state or {}).get(str(car_no))
        s1, s2, s3, s1_ms, s2_ms, s3_ms = sector_times_from_live(live_row, sector_cache)
        has_live = bool(
            live_row
            and (
                best_lap_ms > 0
                or last_lap_ms > 0
                or bool(str(live_row.get("LAPS", "")).strip())
            )
        )

        rows.append(
            TimingRow(
                pos=str(live_row.get("POS", "")).strip() or "-",
                pic=str(live_row.get("PIC", "")).strip() or "-",
                car_no=str(car_no),
                car_class=car_class or "-",
                laps=str(live_row.get("LAPS", "")).strip() or "-",
                current_driver=current_driver or "-",
                all_drivers=get_all_driver_names(entry, use_english) or "-",
                team_name=str(entry.get("TeamNameL", "")).strip() or "-",
                car_name=str(entry.get("CarNameL", "")).strip() or "-",
                best_lap=ms_to_laptime(best_lap_ms),
                last_lap=ms_to_laptime(last_lap_ms),
                s1=s1,
                s2=s2,
                s3=s3,
                best_lap_ms=best_lap_ms,
                last_lap_ms=last_lap_ms,
                s1_ms=s1_ms,
                s2_ms=s2_ms,
                s3_ms=s3_ms,
                has_live_data=has_live,
            )
        )

    def sort_key(row: TimingRow) -> tuple:
        pic = int(row.pic) if row.pic.isdigit() else 9999
        pos = int(row.pos) if row.pos.isdigit() else 9999
        return (pic, pos, int(row.car_no) if row.car_no.isdigit() else 9999)

    rows.sort(key=sort_key)
    return rows


def build_driver_rows(
    master: dict[str, Any],
    live: dict[str, Any],
    class_filter: str | None,
    use_english: bool = False,
    sector_state: dict[str, Any] | None = None,
) -> list[DriverRow]:
    entry_info: dict[str, Any] = master.get("EntryInfo", {})
    live_data: list[dict[str, Any]] = live.get("LiveData", [])
    live_by_car = {str(row.get("CarNo", "")): row for row in live_data}
    driver_rows: list[DriverRow] = []

    for car_no, entry in entry_info.items():
        car_class = str(entry.get("ClassStr", "")).strip()
        if class_filter and car_class != class_filter:
            continue

        live_row = live_by_car.get(str(car_no), {})
        current_driver_id = str(live_row.get("Driver", "")).strip()
        best_lap_ms = int(live_row.get("BestLapTime", 0) or 0)
        last_lap_ms = int(live_row.get("LapTime", 0) or 0)
        sector_cache = (sector_state or {}).get(str(car_no))
        s1, s2, s3, s1_ms, s2_ms, s3_ms = sector_times_from_live(live_row, sector_cache)

        for slot in sorted(entry.get("Drivers", {}).keys()):
            driver_name = get_driver_name(entry, slot, use_english)
            if not driver_name:
                continue

            driver_rows.append(
                DriverRow(
                    driver_name=driver_name,
                    driver_slot=slot,
                    car_no=str(car_no),
                    car_class=car_class or "-",
                    team_name=str(entry.get("TeamNameL", "")).strip() or "-",
                    car_name=str(entry.get("CarNameL", "")).strip() or "-",
                    best_lap=ms_to_laptime(best_lap_ms),
                    last_lap=ms_to_laptime(last_lap_ms),
                    s1=s1,
                    s2=s2,
                    s3=s3,
                    best_lap_ms=best_lap_ms,
                    last_lap_ms=last_lap_ms,
                    s1_ms=s1_ms,
                    s2_ms=s2_ms,
                    s3_ms=s3_ms,
                    is_current=slot == current_driver_id and bool(current_driver_id),
                    pos=str(live_row.get("POS", "")).strip() or "-",
                    pic=str(live_row.get("PIC", "")).strip() or "-",
                    laps=str(live_row.get("LAPS", "")).strip() or "-",
                )
            )

    driver_rows.sort(key=lambda row: (row.driver_name, row.car_no))
    return driver_rows


def search_rows(rows: list[TimingRow], query: str) -> list[TimingRow]:
    query = query.strip().casefold()
    if not query:
        return rows

    matched: list[TimingRow] = []
    for row in rows:
        haystack = " ".join(
            [row.car_no, row.current_driver, row.all_drivers, row.team_name, row.car_name]
        ).casefold()
        if query in haystack:
            matched.append(row)
    return matched


def format_update_time(raw: str) -> str:
    raw = str(raw).strip()
    if len(raw) < 15:
        return raw or "-"

    try:
        dt = datetime.strptime(raw[:15], "%Y%m%d_%H%M%S")
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return raw


def race_status_text(live: dict[str, Any]) -> str:
    live_data = live.get("LiveData", [])
    if has_active_timing(live_data):
        return "レースデータ受信中"
    if int(live.get("IsFinal", 0) or 0) == 1:
        return "待機中（エントリー情報のみ / タイム未配信）"
    return "待機中（タイム未配信）"


def clear_screen() -> None:
    if sys.platform == "win32":
        import os

        os.system("cls")
    else:
        print("\033[2J\033[H", end="")


def print_table(rows: list[TimingRow], header_lines: list[str]) -> None:
    for line in header_lines:
        print(line)

    if not rows:
        print("\n該当データがありません。")
        return

    columns = [
        ("POS", 4),
        ("PIC", 4),
        ("No.", 5),
        ("Class", 7),
        ("Driver", 18),
        ("BestLap", 12),
        ("LastLap", 12),
        ("S1", 10),
        ("S2", 10),
        ("S3", 10),
        ("Laps", 5),
        ("Team", 24),
    ]

    header = " ".join(name.ljust(width) for name, width in columns)
    print()
    print(header)
    print("-" * len(header))

    for row in rows:
        values = [
            row.pos,
            row.pic,
            row.car_no,
            row.car_class,
            row.current_driver[:18],
            row.best_lap,
            row.last_lap,
            row.s1,
            row.s2,
            row.s3,
            row.laps,
            row.team_name[:24],
        ]
        print(" ".join(value.ljust(width) for value, (_, width) in zip(values, columns)))

    print()
    print(f"表示件数: {len(rows)}")


def print_search_results(rows: list[TimingRow], query: str) -> None:
    print(f"\n検索: \"{query}\"")
    if not rows:
        print("該当するドライバーが見つかりませんでした。")
        return

    for row in rows:
        print("-" * 60)
        print(f"車番      : {row.car_no}")
        print(f"クラス    : {row.car_class}")
        print(f"現在走行  : {row.current_driver}")
        print(f"全ドライバー: {row.all_drivers}")
        print(f"BestLap   : {row.best_lap}")
        print(f"LastLap   : {row.last_lap}")
        print(f"チーム    : {row.team_name}")
        print(f"マシン    : {row.car_name}")


def load_timing_data(
    session: requests.Session,
    class_filter: str | None,
    use_english: bool,
) -> tuple[list[TimingRow], dict[str, Any], dict[str, Any]]:
    master = fetch_json(session, "master.json")
    live = fetch_json(session, "livemoni.json")
    rows = build_rows(master, live, class_filter, use_english)
    return rows, master, live


def ensure_output_dir(output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def build_snapshot(
    rows: list[TimingRow],
    master: dict[str, Any],
    live: dict[str, Any],
) -> dict[str, Any]:
    return {
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "race_name": master.get("RaceNameL", ""),
        "race_year": master.get("RaceYear", ""),
        "round_no": master.get("RoundNo", ""),
        "update_time": live.get("UpdateTime", ""),
        "update_time_display": format_update_time(live.get("UpdateTime", "")),
        "status": race_status_text(live),
        "row_count": len(rows),
        "rows": [asdict(row) for row in rows],
    }


def save_latest_files(
    output_dir: Path,
    rows: list[TimingRow],
    master: dict[str, Any],
    live: dict[str, Any],
) -> dict[str, Path]:
    ensure_output_dir(output_dir)
    snapshot = build_snapshot(rows, master, live)

    json_path = latest_json_path(output_dir)
    csv_path = latest_csv_path(output_dir)

    json_path.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    fieldnames = [
        "saved_at",
        "update_time_display",
        "status",
        "pos",
        "pic",
        "car_no",
        "car_class",
        "current_driver",
        "all_drivers",
        "best_lap",
        "last_lap",
        "s1",
        "s2",
        "s3",
        "best_lap_ms",
        "last_lap_ms",
        "s1_ms",
        "s2_ms",
        "s3_ms",
        "laps",
        "team_name",
        "car_name",
        "has_live_data",
    ]

    with csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "saved_at": snapshot["saved_at"],
                    "update_time_display": snapshot["update_time_display"],
                    "status": snapshot["status"],
                    **asdict(row),
                }
            )

    return {"json": json_path, "csv": csv_path}


def append_history(
    output_dir: Path,
    rows: list[TimingRow],
    master: dict[str, Any],
    live: dict[str, Any],
) -> Path:
    ensure_output_dir(output_dir)
    history_path = history_csv_path(output_dir)
    snapshot = build_snapshot(rows, master, live)

    fieldnames = [
        "saved_at",
        "update_time_display",
        "status",
        "pos",
        "pic",
        "car_no",
        "car_class",
        "current_driver",
        "all_drivers",
        "best_lap",
        "last_lap",
        "s1",
        "s2",
        "s3",
        "best_lap_ms",
        "last_lap_ms",
        "s1_ms",
        "s2_ms",
        "s3_ms",
        "laps",
        "team_name",
        "car_name",
        "has_live_data",
    ]

    write_header = not history_path.exists()
    with history_path.open("a", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "saved_at": snapshot["saved_at"],
                    "update_time_display": snapshot["update_time_display"],
                    "status": snapshot["status"],
                    **asdict(row),
                }
            )

    return history_path

def append_raw_snapshot(
    output_dir: Path,
    master: dict[str, Any],
    live: dict[str, Any],
) -> Path:
    ensure_output_dir(warehouse_dir(output_dir))
    snapshot_path = raw_snapshots_csv_path(output_dir)
    saved_at = datetime.now().isoformat(timespec="seconds")
    fieldnames = [
        "saved_at",
        "update_time_display",
        "status",
        "race_name",
        "race_year",
        "master_json",
        "live_json",
    ]
    row = {
        "saved_at": saved_at,
        "update_time_display": format_update_time(live.get("UpdateTime", "")),
        "status": race_status_text(live),
        "race_name": str(master.get("RaceNameL", "")),
        "race_year": str(master.get("RaceYear", "")),
        "master_json": json.dumps(master, ensure_ascii=False, separators=(",", ":")),
        "live_json": json.dumps(live, ensure_ascii=False, separators=(",", ":")),
    }
    write_header = not snapshot_path.exists()
    with snapshot_path.open("a", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        writer.writerow(row)
    return snapshot_path


def save_drivers_running_csv(
    output_dir: Path,
    driver_rows: list[DriverRow],
    master: dict[str, Any],
    live: dict[str, Any],
) -> Path:
    ensure_output_dir(warehouse_dir(output_dir))
    path = drivers_running_csv_path(output_dir)
    snapshot = build_snapshot([], master, live)
    running_rows = [row for row in driver_rows if row.is_current]
    fieldnames = [
        "saved_at",
        "update_time_display",
        "status",
        "driver_name",
        "driver_slot",
        "car_no",
        "car_class",
        "team_name",
        "car_name",
        "best_lap",
        "last_lap",
        "s1",
        "s2",
        "s3",
        "best_lap_ms",
        "last_lap_ms",
        "s1_ms",
        "s2_ms",
        "s3_ms",
        "laps",
        "pos",
        "pic",
        "is_current",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for row in running_rows:
            writer.writerow(
                {
                    "saved_at": snapshot["saved_at"],
                    "update_time_display": snapshot["update_time_display"],
                    "status": snapshot["status"],
                    **asdict(row),
                }
            )
    return path


def _write_sheet_rows(
    worksheet: Any,
    headers: list[str],
    rows: list[list[Any]],
    *,
    highlight_rows: set[int] | None = None,
) -> None:
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True)
    team_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = header_font
    for row_index, row in enumerate(rows, start=2):
        worksheet.append(row)
        if highlight_rows and row_index in highlight_rows:
            for cell in worksheet[row_index]:
                cell.fill = team_fill


def export_live_xlsx(
    output_dir: Path,
    rows: list[TimingRow],
    driver_rows: list[DriverRow],
    master: dict[str, Any],
    live: dict[str, Any],
    default_class: str,
    interval: int,
) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl が必要です: pip install openpyxl") from exc

    ensure_output_dir(output_dir)
    path = live_xlsx_path(output_dir)
    temp_path = path.with_suffix(".xlsx.tmp")
    snapshot = build_snapshot(rows, master, live)
    running_rows = [row for row in driver_rows if row.is_current]

    workbook = Workbook()
    settings_sheet = workbook.active
    settings_sheet.title = "設定"
    settings_sheet.append(["項目", "値"])
    for label, value in [
        ("レース名", snapshot["race_name"]),
        ("開催年", snapshot["race_year"]),
        ("ラウンド", snapshot.get("round_no", "")),
        ("更新時刻", snapshot["update_time_display"]),
        ("保存時刻", snapshot["saved_at"]),
        ("状態", snapshot["status"]),
        ("表示クラス", default_class),
        ("取得間隔(秒)", interval),
        ("全車両数", len(rows)),
        ("走行中ドライバー数", len(running_rows)),
        ("生JSON倉庫", str(raw_snapshots_csv_path(output_dir))),
        ("加工済み履歴", str(history_csv_path(output_dir))),
    ]:
        settings_sheet.append([label, value])

    raw_sheet = workbook.create_sheet("元データ")
    raw_headers = [
        "POS", "PIC", "車番", "クラス", "走行中ドライバー", "全ドライバー",
        "BestLap", "LastLap", "S1", "S2", "S3", "周回数", "チーム", "マシン", "ライブデータ",
    ]
    raw_rows: list[list[Any]] = []
    highlight_rows: set[int] = set()
    for index, row in enumerate(rows, start=2):
        if row.car_no in TEAM_CAR_NOS:
            highlight_rows.add(index)
        raw_rows.append([
            row.pos, row.pic, row.car_no, row.car_class, row.current_driver,
            row.all_drivers, row.best_lap, row.last_lap, row.s1, row.s2, row.s3, row.laps,
            row.team_name, row.car_name, "あり" if row.has_live_data else "なし",
        ])
    _write_sheet_rows(raw_sheet, raw_headers, raw_rows, highlight_rows=highlight_rows)

    drivers_sheet = workbook.create_sheet("ドライバー")
    driver_headers = [
        "ドライバー", "スロット", "状態", "車番", "クラス", "BestLap", "LastLap",
        "S1", "S2", "S3", "周回数", "POS", "PIC", "チーム", "マシン",
    ]
    driver_data: list[list[Any]] = []
    driver_highlight: set[int] = set()
    for index, row in enumerate(running_rows, start=2):
        if row.car_no in TEAM_CAR_NOS:
            driver_highlight.add(index)
        driver_data.append([
            row.driver_name, row.driver_slot, "走行中", row.car_no, row.car_class,
            row.best_lap, row.last_lap, row.s1, row.s2, row.s3, row.laps, row.pos, row.pic,
            row.team_name, row.car_name,
        ])
    _write_sheet_rows(drivers_sheet, driver_headers, driver_data, highlight_rows=driver_highlight)

    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 48)

    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except OSError:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        return None
    return path


def dashboard_theme_css(*, drivers_page: bool = False, index_page: bool = False) -> str:
    sticky_panel_bg = "#3a2f36" if drivers_page else ""
    sticky_live_bg = "#46343b" if drivers_page else ""
    sticky_rules = ""
    if drivers_page:
        sticky_rules = f"""
    #tableWrap {{
      max-height: calc(100vh - 220px);
    }}
    .sticky-col {{
      position: sticky;
      background: {sticky_panel_bg};
    }}
    tr.row-live .sticky-col {{ background: {sticky_live_bg}; }}
    tr.row-team .sticky-col {{ background: #523a42; }}
    thead .sticky-col {{
      background: #5a3640;
      z-index: 4;
    }}
    .sticky-col-1 {{
      left: 0;
      min-width: 160px;
      max-width: 160px;
    }}
    .sticky-col-2 {{
      left: 160px;
      min-width: 64px;
      max-width: 64px;
    }}
    .sticky-col-3 {{
      left: 224px;
      min-width: 88px;
      max-width: 88px;
      box-shadow: 4px 0 8px rgba(0, 0, 0, 0.35);
    }}"""

    return f"""
    :root {{
      --bg: #2c242a;
      --panel: #382e34;
      --panel-2: #43363e;
      --border: #5e4650;
      --text: #ffffff;
      --muted: #e3d4d8;
      --text-strong: #ffffff;
      --text-driver: #ffffff;
      --text-driver-live: #fff4c8;
      --text-body: #ffffff;
      --text-faint: #ecdfe2;
      --text-lap: #ffffff;
      --text-slot: #cfe2ff;
      --accent: #e10600;
      --accent-soft: #ff4d5a;
      --accent-glow: rgba(225, 6, 0, 0.28);
      --class: #ff8a8a;
      --best: #fff0c8;
      --live: rgba(225, 6, 0, 0.14);
      --team: rgba(225, 6, 0, 0.22);
      --page-font: calc(0.92rem + 3pt);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "Hiragino Sans", "Yu Gothic UI", sans-serif;
      background:
        radial-gradient(circle at top right, rgba(225, 6, 0, 0.12), transparent 28%),
        linear-gradient(180deg, #352a32 0%, var(--bg) 100%);
      color: var(--text);
      min-height: 100vh;
    }}
    .wrap {{ max-width: 1600px; margin: 0 auto; padding: 20px; }}
    .wrap, .wrap * {{ font-size: var(--page-font) !important; }}
    .header.hero {{
      position: relative;
      overflow: hidden;
      border: 1px solid var(--border);
      border-radius: 12px;
      margin-bottom: 12px;
      background: var(--panel);
    }}
    .hero-banner {{
      position: absolute;
      inset: 0;
      background-image: url("{TEAM_BANNER_ASSET}");
      background-size: cover;
      background-position: center;
      opacity: 0.34;
    }}
    .hero-banner::after {{
      content: "";
      position: absolute;
      inset: 0;
      background: linear-gradient(90deg, rgba(44, 36, 42, 0.94) 0%, rgba(44, 36, 42, 0.68) 45%, rgba(44, 36, 42, 0.86) 100%);
    }}
    .hero-inner {{
      position: relative;
      z-index: 1;
      display: flex;
      gap: 14px;
      align-items: center;
      padding: 10px 16px;
      flex-wrap: nowrap;
    }}
    .team-logo {{
      width: 52px;
      height: 52px;
      border-radius: 10px;
      border: 2px solid rgba(255, 255, 255, 0.14);
      box-shadow: 0 6px 18px rgba(0, 0, 0, 0.35);
      object-fit: cover;
      background: #fff;
      flex: 0 0 auto;
    }}
    .hero-main {{
      flex: 1 1 0;
      min-width: 0;
      display: flex;
      align-items: center;
      gap: 14px 18px;
      flex-wrap: wrap;
    }}
    .hero-head {{
      display: flex;
      align-items: baseline;
      gap: 8px 12px;
      flex-wrap: wrap;
      flex: 0 1 auto;
      min-width: 0;
    }}
    .hero-head h1 {{
      margin: 0;
      font-size: 1.2rem;
      line-height: 1.2;
      white-space: nowrap;
    }}
    .eyebrow {{
      margin: 0;
      color: #ff9f9f;
      font-size: 0.72rem;
      font-weight: 700;
      letter-spacing: 0.06em;
      text-transform: uppercase;
      white-space: nowrap;
    }}
    .tagline {{
      margin: 0;
      color: var(--muted);
      font-size: 0.8rem;
      white-space: nowrap;
    }}
    .hero-meta-row {{
      display: flex;
      align-items: center;
      gap: 8px 12px;
      flex: 1 1 320px;
      flex-wrap: wrap;
      justify-content: flex-end;
      min-width: 0;
    }}
    .meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 4px 12px;
      color: var(--muted);
      font-size: 0.78rem;
      line-height: 1.3;
    }}
    .meta-item {{
      white-space: nowrap;
    }}
    .meta-item:not(:last-child)::after {{
      content: "|";
      margin-left: 12px;
      color: rgba(255, 255, 255, 0.18);
    }}
    .status {{
      display: inline-block;
      padding: 3px 9px;
      border-radius: 999px;
      font-size: 0.76rem;
      font-weight: 700;
      white-space: nowrap;
      flex: 0 0 auto;
    }}
    .status-live {{ background: rgba(225, 6, 0, 0.22); color: #ffb3b3; }}
    .status-wait {{ background: rgba(255, 180, 0, 0.15); color: #ffd166; }}
    .x-link {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      color: #ffb4b4;
      text-decoration: none;
      font-size: 0.78rem;
      font-weight: 700;
      white-space: nowrap;
      flex: 0 0 auto;
    }}
    .x-link:hover {{ color: #fff; }}
    .toolbar {{
      display: flex;
      gap: 12px;
      align-items: center;
      margin-bottom: 16px;
      flex-wrap: wrap;
    }}
    .toolbar select, .toolbar input {{
      padding: 10px 14px;
      border-radius: 8px;
      border: 1px solid var(--border);
      background: var(--panel);
      color: var(--text);
      font-size: calc(1rem + 2pt);
    }}
    .toolbar select {{ min-width: 140px; }}
    .toolbar input {{ flex: 1; min-width: 220px; }}
    .toolbar .info {{ color: var(--muted); font-size: calc(0.9rem + 2pt); }}
    .nav-link {{
      color: #ffb4b4;
      text-decoration: none;
      padding: 10px 14px;
      border: 1px solid var(--border);
      border-radius: 8px;
      white-space: nowrap;
      background: rgba(225, 6, 0, 0.08);
    }}
    .nav-link:hover {{ background: rgba(225, 6, 0, 0.18); }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--border);
      border-radius: 12px;
      background: var(--panel);
    }}
    table {{
      width: 100%;
      border-collapse: {"separate" if drivers_page else "collapse"};
      border-spacing: 0;
      min-width: {"1000px" if drivers_page else "1100px"};
      font-size: 0.92rem;
    }}
    th, td {{
      padding: 10px 12px;
      border-bottom: 1px solid var(--border);
      text-align: left;
      vertical-align: top;
    }}
    tbody td {{ font-size: calc(1em + 3pt); }}
    th {{
      position: sticky;
      top: 0;
      background: #5a3640;
      color: #ffd7d7;
      font-size: calc(0.8rem + 2pt);
      letter-spacing: 0.04em;
      text-transform: uppercase;
      z-index: 2;
    }}
    tr:hover {{ background: rgba(255, 255, 255, 0.03); }}
    tr.row-live {{ background: var(--live); }}
    tr.row-team {{ background: var(--team); box-shadow: inset 3px 0 0 var(--accent); }}
    .num {{ text-align: center; white-space: nowrap; color: var(--text-body); }}
    .car-no {{ font-weight: 800; color: var(--text-strong); }}
    .class-badge {{
      display: inline-block;
      background: rgba(225, 6, 0, 0.18);
      color: var(--class);
      padding: 2px 8px;
      border-radius: 6px;
      font-weight: 700;
      font-size: 0.82rem;
    }}
    .driver {{
      font-weight: 800;
      min-width: 120px;
      color: var(--text-driver);
      font-size: 1.02em;
      letter-spacing: 0.02em;
    }}
    tr.row-live .driver {{
      color: var(--text-driver-live);
      text-shadow: 0 0 14px rgba(255, 244, 200, 0.28);
    }}
    .driver-slot {{ color: var(--text-slot); font-weight: 700; }}
    .drivers {{ color: var(--muted); min-width: 160px; font-size: 0.9em; }}
    .team, .car {{ color: var(--text-faint); min-width: 160px; font-size: 0.88em; }}
    .lap {{
      font-family: Consolas, "Courier New", monospace;
      white-space: nowrap;
      text-align: right;
      min-width: 90px;
      color: var(--text-lap);
    }}
    .sector {{
      min-width: 72px;
      font-size: 0.9em;
    }}
    .lap-last {{ color: var(--text-body); }}
    .lap-best {{
      color: var(--best);
      font-weight: 800;
      text-shadow: 0 0 10px rgba(255, 240, 200, 0.35);
    }}
    .current-badge {{
      display: inline-block;
      background: rgba(225, 6, 0, 0.28);
      color: #ffe0e0;
      padding: 2px 8px;
      border-radius: 6px;
      font-size: 0.82rem;
      font-weight: 700;
    }}
    .hidden {{ display: none; }}
    .lap-panel {{
      margin-top: 16px;
      border: 1px solid var(--border);
      border-radius: 12px;
      background: var(--panel-2);
      padding: 16px 18px;
    }}
    .lap-panel h2 {{
      margin: 0 0 12px;
      font-size: 1.1rem;
    }}
    .lap-panel .summary {{
      color: var(--muted);
      font-size: 0.9rem;
      margin-bottom: 12px;
    }}
    .lap-panel .table-wrap {{
      max-height: 420px;
    }}
    .lap-panel table th {{
      font-size: calc(0.8rem + 7pt);
    }}
    .lap-panel table tbody td {{
      font-size: calc(1em + 6pt);
    }}
    .lap-panel th.sortable {{
      cursor: pointer;
      user-select: none;
    }}
    .lap-panel th.sortable:hover {{
      color: #fff;
    }}
    .lap-panel th.sortable .sort-icon {{
      margin-left: 4px;
      font-size: 0.7rem;
      opacity: 0.75;
    }}
    .lap-panel th.sortable.sort-asc .sort-icon::after {{ content: "▲"; }}
    .lap-panel th.sortable.sort-desc .sort-icon::after {{ content: "▼"; }}
    .footer {{
      margin-top: 14px;
      color: var(--muted);
      font-size: 0.85rem;
      line-height: 1.6;
    }}{"" if not index_page else f"""
    .wrap.index-layout {{
      max-width: 1920px;
    }}
    .main-split {{
      display: flex;
      gap: 16px;
      align-items: flex-start;
    }}
    .main-split.layout-horizontal {{
      flex-direction: row;
    }}
    .main-split.layout-vertical {{
      flex-direction: column;
    }}
    .main-split.layout-vertical .data-panel {{
      order: 1;
    }}
    .main-split.layout-vertical .video-panel {{
      order: 2;
    }}
    .video-panel {{
      min-width: 0;
    }}
    .video-controls {{
      display: flex;
      justify-content: center;
      align-items: center;
      gap: 10px;
      margin-bottom: 8px;
      color: var(--muted);
      font-size: 0.85rem;
    }}
    .video-controls input[type="range"] {{
      width: min(360px, 42vw);
      accent-color: var(--accent);
    }}
    .video-size-value {{
      min-width: 56px;
      color: var(--text-body);
      font-variant-numeric: tabular-nums;
      text-align: right;
    }}
    .youtube-settings {{
      margin-top: 16px;
      border: 1px solid var(--border);
      border-radius: 12px;
      background: var(--panel);
      padding: 12px;
    }}
    .youtube-settings h2 {{
      margin: 0 0 8px;
      color: var(--text-strong);
    }}
    .youtube-settings-row {{
      display: flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
    }}
    .youtube-settings input {{
      flex: 1 1 420px;
      min-width: 240px;
      padding: 10px 12px;
      border: 1px solid var(--border);
      border-radius: 8px;
      background: var(--panel-2);
      color: var(--text);
    }}
    .youtube-settings button {{
      border: 1px solid var(--border);
      border-radius: 8px;
      background: rgba(225, 6, 0, 0.12);
      color: #ffdddd;
      padding: 10px 14px;
      cursor: pointer;
    }}
    .youtube-settings button:hover {{
      background: rgba(225, 6, 0, 0.22);
      color: #fff;
    }}
    .youtube-settings .summary {{
      margin-top: 8px;
      color: var(--muted);
    }}
    .main-split.layout-horizontal .video-panel {{
      flex: 0 0 48%;
      min-width: 300px;
      position: sticky;
      top: 16px;
    }}
    .main-split.layout-vertical .video-panel {{
      flex: none;
      width: 100%;
      position: static;
    }}
    .video-frame {{
      position: relative;
      width: 100%;
      padding-bottom: 56.25%;
      border: 1px solid var(--border);
      border-radius: 12px;
      overflow: hidden;
      background: #000;
    }}
    .main-split.layout-vertical .video-frame {{
      max-width: var(--video-max-width, 960px);
      margin: 0 auto;
    }}
    .video-frame iframe {{
      position: absolute;
      inset: 0;
      width: 100%;
      height: 100%;
      border: 0;
    }}
    .data-panel {{
      flex: 1 1 0;
      min-width: 0;
      width: 100%;
    }}
    .page-split {{
      display: flex;
      gap: 16px;
      align-items: flex-start;
    }}
    .page-main {{
      flex: 1 1 0;
      min-width: 0;
    }}
    .cars-panel {{
      width: 100%;
      min-width: 0;
    }}
    .cars-panel table tbody td:nth-child(1) {{ color: #ffd54a; }}
    .cars-panel table tbody td:nth-child(2) {{ color: #b9a3ff; }}
    .cars-panel table tbody td:nth-child(3) {{ color: #6fe3ff; }}
    .cars-panel table tbody td:nth-child(5) {{ color: #ffb067; }}
    .cars-panel table tbody td:nth-child(6) {{ color: #a7d98f; }}
    .cars-panel table tbody td:nth-child(8),
    .cars-panel table tbody td:nth-child(9),
    .cars-panel table tbody td:nth-child(10),
    .cars-panel table tbody td:nth-child(11) {{ color: #ffffff; }}
    .cars-panel table tbody td:nth-child(12) {{ color: #7fb4ff; }}
    .cars-panel table tbody td:nth-child(13) {{ color: #f4a3c4; }}
    .cars-panel table tbody td:nth-child(14) {{ color: #6fe0c8; }}
    .main-split.layout-horizontal .data-panel #tableWrap,
    .main-split.layout-vertical .data-panel #tableWrap {{
      max-height: calc(100vh - 240px);
    }}
    .lap-column {{
      position: fixed;
      top: 16px;
      right: 16px;
      width: var(--lap-col-width, 260px);
      border: 1px solid var(--border);
      border-radius: 12px;
      background: var(--panel-2);
      display: flex;
      flex-direction: column;
      height: calc(100vh - 32px);
      max-height: calc(100vh - 32px);
      z-index: 50;
      box-shadow: 0 14px 44px rgba(0, 0, 0, 0.5);
    }}
    .lap-column.dragging {{
      user-select: none;
      box-shadow: 0 18px 56px rgba(0, 0, 0, 0.6);
    }}
    @media (max-width: 1100px) {{
      .lap-column {{
        width: var(--lap-col-width, min(90vw, 360px));
        max-height: calc(100vh - 120px);
      }}
    }}
    .lap-column.hidden {{
      display: none;
    }}
    .lap-column-header {{
      padding: 8px 12px 8px;
      border-bottom: 1px solid var(--border);
      flex: 0 0 auto;
      cursor: move;
    }}
    .lap-col-bar {{
      display: flex;
      align-items: center;
      gap: 8px;
      justify-content: space-between;
    }}
    .lap-col-tools {{
      display: flex;
      align-items: center;
      gap: 6px;
      flex: 0 0 auto;
    }}
    .lap-col-tools input[type="range"] {{
      width: 84px;
      accent-color: var(--accent);
      cursor: pointer;
    }}
    .lap-col-reset {{
      border: 1px solid var(--border);
      border-radius: 6px;
      background: transparent;
      color: #ffb4b4;
      font-size: 0.78rem;
      line-height: 1;
      padding: 3px 7px;
      cursor: pointer;
    }}
    .lap-col-reset:hover {{
      background: rgba(225, 6, 0, 0.16);
      color: #fff;
    }}
    .lap-group-btn.hidden {{
      display: none;
    }}
    .lap-group-btn.active {{
      background: rgba(225, 6, 0, 0.3);
      border-color: var(--accent-soft);
      color: #fff;
      font-weight: 700;
    }}
    .lap-group-row td {{
      background: #4a3a42;
      color: var(--text-driver);
      font-weight: 800;
      letter-spacing: 0.02em;
      border-bottom: 1px solid var(--border);
      position: sticky;
      top: 0;
      z-index: 1;
    }}
    .lap-column h2 {{
      margin: 0;
      font-size: 0.92rem;
      line-height: 1.3;
      flex: 1 1 auto;
      min-width: 0;
    }}
    .lap-column .summary {{
      color: var(--muted);
      font-size: 0.76rem;
      margin: 0;
    }}
    .lap-filter-pills {{
      display: flex;
      flex-wrap: wrap;
      gap: 5px;
      margin-top: 8px;
    }}
    .lap-filter-pills button {{
      border: 1px solid var(--border);
      border-radius: 999px;
      background: transparent;
      color: var(--muted);
      padding: 4px 9px;
      font-size: 0.72rem;
      line-height: 1;
      cursor: pointer;
    }}
    .lap-filter-pills button.active {{
      background: rgba(225, 6, 0, 0.28);
      border-color: var(--accent-soft);
      color: #fff;
      font-weight: 700;
    }}
    .lap-column th.sortable {{
      cursor: pointer;
      user-select: none;
    }}
    .lap-column th.sortable:hover {{
      color: #fff;
    }}
    .lap-column th.sortable .sort-icon {{
      margin-left: 4px;
      font-size: 0.7rem;
      opacity: 0.75;
    }}
    .lap-column th.sortable.sort-asc .sort-icon::after {{ content: "▲"; }}
    .lap-column th.sortable.sort-desc .sort-icon::after {{ content: "▼"; }}
    .lap-column .table-wrap {{
      flex: 1 1 auto;
      overflow: auto;
      border: none;
      border-radius: 0 0 12px 12px;
      max-height: none;
    }}
    .lap-column table {{
      min-width: 0;
      font-size: calc(0.8rem + 2pt);
    }}
    .lap-column th,
    .lap-column td {{
      padding: 6px 8px;
    }}
    .lap-column table tbody td {{
      font-size: calc(0.92rem + 3pt);
    }}
    .driver-filter {{
      border: 1px solid var(--border);
      border-radius: 10px;
      background: var(--panel);
      padding: 8px 12px 10px;
      min-width: 240px;
      max-width: 460px;
      flex: 1 1 320px;
    }}
    .driver-filter-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      margin-bottom: 8px;
      color: var(--muted);
      font-size: 0.74rem;
      font-weight: 700;
      letter-spacing: 0.02em;
      text-transform: uppercase;
    }}
    .driver-clear-btn {{
      border: 1px solid var(--border);
      border-radius: 999px;
      background: transparent;
      color: #ffb4b4;
      font-size: 0.72rem;
      padding: 3px 12px;
      cursor: pointer;
      transition: background 0.15s ease, color 0.15s ease;
    }}
    .driver-clear-btn:hover {{
      background: rgba(225, 6, 0, 0.16);
      color: #fff;
    }}
    .driver-checkboxes {{
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      max-height: 96px;
      overflow-y: auto;
      padding-right: 2px;
    }}
    .driver-checkboxes::-webkit-scrollbar {{
      width: 8px;
    }}
    .driver-checkboxes::-webkit-scrollbar-thumb {{
      background: rgba(255, 255, 255, 0.12);
      border-radius: 999px;
    }}
    .driver-checkboxes label {{
      display: inline-flex;
      align-items: center;
      gap: 0;
      font-size: 0.82rem;
      line-height: 1;
      white-space: nowrap;
      cursor: pointer;
      color: var(--text-body);
      padding: 6px 12px;
      border: 1px solid var(--border);
      border-radius: 999px;
      background: var(--panel-2);
      transition: background 0.15s ease, border-color 0.15s ease, color 0.15s ease;
      user-select: none;
    }}
    .driver-checkboxes label:hover {{
      border-color: var(--accent-soft);
      color: var(--text-strong);
    }}
    .driver-checkboxes label:has(input:checked) {{
      background: rgba(225, 6, 0, 0.28);
      border-color: var(--accent-soft);
      color: #fff;
      font-weight: 700;
      box-shadow: 0 0 10px rgba(225, 6, 0, 0.22);
    }}
    .driver-checkboxes input {{
      position: absolute;
      opacity: 0;
      width: 0;
      height: 0;
      pointer-events: none;
    }}
    .driver-pickable {{
      cursor: pointer;
      text-decoration: underline dotted;
      text-decoration-color: rgba(255, 216, 154, 0.45);
      text-underline-offset: 3px;
    }}
    .driver-pickable:hover {{
      color: var(--text-driver-live);
    }}"""}{sticky_rules}"""


def dashboard_header_html(page_title: str) -> str:
    return f"""
    <div class="header hero">
      <div class="hero-banner"></div>
      <div class="hero-inner">
        <img class="team-logo" src="{html.escape(TEAM_LOGO_ASSET)}" alt="{html.escape(TEAM_NAME)}">
        <div class="hero-main">
          <div class="hero-head">
            <p class="eyebrow">#821 / #822 ST-5F</p>
            <h1>{html.escape(page_title)}</h1>
            <p class="tagline">{html.escape(TEAM_TAGLINE)}</p>
          </div>
          <div class="hero-meta-row">
            <div class="meta" id="metaBox">読み込み中...</div>
            <span class="status status-wait" id="statusBadge">更新待ち</span>
            <a class="x-link" href="{html.escape(TEAM_X_URL)}" target="_blank" rel="noopener">@AndLegal_Racing</a>
          </div>
        </div>
      </div>
    </div>"""


def embed_view_data_script(view_data: dict[str, Any] | None) -> str:
    if not view_data:
        return ""
    payload = json.dumps(view_data, ensure_ascii=False)
    return f"<script>window.__INITIAL_VIEW_DATA__ = {payload};</script>\n"


def start_static_server(root_dir: Path, port: int) -> http.server.ThreadingHTTPServer:
    handler = functools.partial(
        http.server.SimpleHTTPRequestHandler,
        directory=str(root_dir),
    )
    server = http.server.ThreadingHTTPServer(("127.0.0.1", port), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server


def dashboard_page_url(port: int, page: str = "data/index.html") -> str:
    return f"http://127.0.0.1:{port}/{page}"


def save_html_file(
    output_dir: Path,
    interval: int,
    default_class: str,
    initial_view_data: dict[str, Any] | None = None,
) -> Path:
    ensure_output_dir(output_dir)
    html_path = output_dir / "index.html"
    data_url = view_data_path(output_dir).name

    page = f"""<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{html.escape(TEAM_NAME)} | ライブタイミング</title>
  <style>{dashboard_theme_css(index_page=True)}</style>
</head>
<body>
  <div class="wrap index-layout">
    {dashboard_header_html(f"{TEAM_NAME} ライブタイミング")}

    <div class="page-split">
      <div class="page-main">
    <div id="mainSplit" class="main-split layout-vertical">
      <div class="video-panel">
        <div class="video-controls">
          <label for="videoSize">YouTubeサイズ</label>
          <input id="videoSize" type="range" min="480" max="1400" step="20" value="960">
          <span id="videoSizeValue" class="video-size-value">960px</span>
        </div>
        <div class="video-frame">
          <iframe
            id="youtubeFrame"
            src="https://www.youtube.com/embed/{html.escape(TEAM_YOUTUBE_VIDEO_ID)}"
            title="YouTube live"
            allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture; web-share"
            allowfullscreen
            referrerpolicy="strict-origin-when-cross-origin"
          ></iframe>
        </div>
      </div>

      <div class="data-panel">
        <div class="toolbar">
          <select id="classFilter"></select>
          <div class="driver-filter" id="driverFilterBox">
            <div class="driver-filter-head">
              <span>ドライバー（複数選択可）</span>
              <button type="button" id="driverClearBtn" class="driver-clear-btn">クリア</button>
            </div>
            <div class="driver-checkboxes" id="driverCheckboxes"></div>
          </div>
          <input id="search" type="search" placeholder="車番・ドライバー・チーム名で検索...">
          <a class="nav-link" href="drivers.html">ドライバー一覧</a>
          <a class="nav-link" href="timing_live.xlsx" download>Excel出力</a>
          <div class="info">表示: <span id="count">0</span> 件 / <span id="pollInfo">{interval}秒ごとに自動更新</span></div>
        </div>

          <div class="cars-panel">
            <div class="table-wrap" id="tableWrap">
              <table>
                <thead>
                  <tr>
                    <th>POS</th>
                    <th>PIC</th>
                    <th>No.</th>
                    <th>Class</th>
                    <th>Driver</th>
                    <th>全ドライバー</th>
                    <th>BestLap</th>
                    <th>LastLap</th>
                    <th>S1</th>
                    <th>S2</th>
                    <th>S3</th>
                    <th>Laps</th>
                    <th>Team</th>
                    <th>Car</th>
                  </tr>
                </thead>
                <tbody id="rows"></tbody>
              </table>
            </div>

            <div class="footer">
              {html.escape(TEAM_NAME)} 専用ダッシュボード / チェックまたは名前ダブルクリックで周回履歴を表示
            </div>
          </div>
        </div>
      </div>
      </div>

      <div id="lapColumn" class="lap-column hidden">
        <div class="lap-column-header" id="lapColumnHandle">
          <div class="lap-col-bar">
            <h2 id="lapPanelTitle">周回タイム</h2>
            <div class="lap-col-tools">
              <button type="button" id="lapGroupBtn" class="lap-col-reset lap-group-btn hidden" title="ドライバーごとにグループ表示">グループ</button>
              <input id="lapColWidth" type="range" min="200" max="680" step="10" value="260" title="カラム幅">
              <button type="button" id="lapColReset" class="lap-col-reset" title="位置と幅をリセット">⤺</button>
            </div>
          </div>
          <div class="summary" id="lapPanelSummary"></div>
          <div id="lapFilterPills" class="lap-filter-pills"></div>
        </div>
        <div class="table-wrap" id="lapTableWrap">
          <table>
            <thead id="lapTableHead">
              <tr>
                <th class="sortable sort-asc" data-lap-sort="lap">Lap<span class="sort-icon"></span></th>
                <th class="sortable" data-lap-sort="time">Time<span class="sort-icon"></span></th>
                <th class="sortable lap-driver-col hidden" data-lap-sort="driver">Driver<span class="sort-icon"></span></th>
                <th class="sortable" data-lap-sort="slot">Slot<span class="sort-icon"></span></th>
              </tr>
            </thead>
            <tbody id="lapRows"></tbody>
          </table>
        </div>
      </div>
  </div>

    <section class="youtube-settings" aria-label="YouTube URL設定">
      <h2>YouTube URL設定</h2>
      <div class="youtube-settings-row">
        <input id="youtubeUrlInput" type="url" placeholder="YouTube URL または動画IDを貼り付け">
        <button type="button" id="youtubeApplyBtn">反映</button>
        <button type="button" id="youtubeResetBtn">既定に戻す</button>
      </div>
      <div class="summary" id="youtubeUrlStatus">URLを貼り付けて反映できます。</div>
    </section>

  {embed_view_data_script(initial_view_data)}
  <script>
    const DATA_URL = "{html.escape(data_url)}";
    const DEFAULT_CLASS = "{html.escape(default_class)}";
    const TEAM_CAR_NOS = {json.dumps(list(TEAM_CAR_NOS))};
    const DEFAULT_YOUTUBE_VIDEO_ID = "{html.escape(TEAM_YOUTUBE_VIDEO_ID)}";
    const CLASS_STORAGE_KEY = "st_selected_class";
    const DRIVER_STORAGE_KEY = "st_index_selected_drivers";
    const LAP_FILTER_STORAGE_KEY = "st_index_lap_driver_filter";
    const LAP_SORT_STORAGE_KEY = "st_index_lap_sort";
    const VIDEO_SIZE_STORAGE_KEY = "st_index_video_size";
    const YOUTUBE_URL_STORAGE_KEY = "st_index_youtube_url";
    const LAP_COL_WIDTH_STORAGE_KEY = "st_index_lap_col_width";
    const LAP_COL_POS_STORAGE_KEY = "st_index_lap_col_pos";
    const LAP_GROUP_STORAGE_KEY = "st_index_lap_group_by_driver";
    const POLL_INTERVAL = {interval} * 1000;
    const classFilter = document.getElementById("classFilter");
    const driverCheckboxes = document.getElementById("driverCheckboxes");
    const driverClearBtn = document.getElementById("driverClearBtn");
    const search = document.getElementById("search");
    const count = document.getElementById("count");
    const rowsBody = document.getElementById("rows");
    const tableWrap = document.getElementById("tableWrap");
    const videoSize = document.getElementById("videoSize");
    const videoSizeValue = document.getElementById("videoSizeValue");
    const videoFrame = document.querySelector(".video-frame");
    const youtubeFrame = document.getElementById("youtubeFrame");
    const youtubeUrlInput = document.getElementById("youtubeUrlInput");
    const youtubeApplyBtn = document.getElementById("youtubeApplyBtn");
    const youtubeResetBtn = document.getElementById("youtubeResetBtn");
    const youtubeUrlStatus = document.getElementById("youtubeUrlStatus");
    const lapColumn = document.getElementById("lapColumn");
    const lapPanelTitle = document.getElementById("lapPanelTitle");
    const lapPanelSummary = document.getElementById("lapPanelSummary");
    const lapFilterPills = document.getElementById("lapFilterPills");
    const lapColumnHandle = document.getElementById("lapColumnHandle");
    const lapColWidth = document.getElementById("lapColWidth");
    const lapColReset = document.getElementById("lapColReset");
    const lapGroupBtn = document.getElementById("lapGroupBtn");
    const lapRowsBody = document.getElementById("lapRows");
    const lapTableWrap = document.getElementById("lapTableWrap");
    const lapDriverCol = document.querySelector(".lap-driver-col");
    const lapSortHeaders = Array.from(document.querySelectorAll("[data-lap-sort]"));
    const metaBox = document.getElementById("metaBox");
    const statusBadge = document.getElementById("statusBadge");
    const mainSplit = document.getElementById("mainSplit");
    let latestData = {{ cars: [], lap_history: [] }};
    let lapSort = loadLapSort();
    let lapDriverFilterSet = loadLapDriverFilter();
    let lapGroupByDriver = localStorage.getItem(LAP_GROUP_STORAGE_KEY) === "1";

    function escapeHtml(value) {{
      return String(value)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");
    }}

    function selectedClass() {{
      return classFilter.value || "ALL";
    }}

    function matchesClass(carClass) {{
      const selected = selectedClass();
      return selected === "ALL" || carClass === selected;
    }}

    function rebuildClassFilter() {{
      const selected = classFilter.value || localStorage.getItem(CLASS_STORAGE_KEY) || latestData.default_class || DEFAULT_CLASS;
      classFilter.innerHTML = '<option value="ALL">全クラス</option>';
      for (const carClass of latestData.classes || []) {{
        const option = document.createElement("option");
        option.value = carClass;
        option.textContent = carClass;
        classFilter.appendChild(option);
      }}
      const values = Array.from(classFilter.options).map((option) => option.value);
      classFilter.value = values.includes(selected) ? selected : "ALL";
    }}

    function loadSelectedDrivers() {{
      try {{
        const raw = localStorage.getItem(DRIVER_STORAGE_KEY);
        if (!raw) return new Set();
        const parsed = JSON.parse(raw);
        if (Array.isArray(parsed)) return new Set(parsed.filter(Boolean));
        if (typeof parsed === "string" && parsed) return new Set([parsed]);
      }} catch (error) {{
        console.warn("ドライバー選択の復元に失敗", error);
      }}
      return new Set();
    }}

    function loadStringSet(storageKey) {{
      try {{
        const raw = localStorage.getItem(storageKey);
        if (!raw) return new Set();
        const parsed = JSON.parse(raw);
        if (Array.isArray(parsed)) return new Set(parsed.filter(Boolean));
      }} catch (error) {{
        console.warn("保存済みフィルタの復元に失敗", error);
      }}
      return new Set();
    }}

    function loadLapDriverFilter() {{
      return loadStringSet(LAP_FILTER_STORAGE_KEY);
    }}

    function saveLapDriverFilter() {{
      if (lapDriverFilterSet.size) {{
        localStorage.setItem(LAP_FILTER_STORAGE_KEY, JSON.stringify(Array.from(lapDriverFilterSet)));
      }} else {{
        localStorage.removeItem(LAP_FILTER_STORAGE_KEY);
      }}
    }}

    function loadLapSort() {{
      try {{
        const parsed = JSON.parse(localStorage.getItem(LAP_SORT_STORAGE_KEY) || "{{}}");
        const key = ["lap", "time", "driver", "slot"].includes(parsed.key) ? parsed.key : "lap";
        const dir = parsed.dir === -1 ? -1 : 1;
        return {{ key, dir }};
      }} catch (error) {{
        console.warn("周回ソートの復元に失敗", error);
        return {{ key: "lap", dir: 1 }};
      }}
    }}

    function saveLapSort() {{
      localStorage.setItem(LAP_SORT_STORAGE_KEY, JSON.stringify(lapSort));
    }}

    function applyVideoSize(rawSize, persist = true) {{
      const size = Math.min(1400, Math.max(480, Number(rawSize) || 960));
      videoFrame.style.setProperty("--video-max-width", `${{size}}px`);
      videoSize.value = String(size);
      videoSizeValue.textContent = `${{size}}px`;
      if (persist) localStorage.setItem(VIDEO_SIZE_STORAGE_KEY, String(size));
    }}

    function extractYoutubeVideoId(value) {{
      const text = String(value || "").trim();
      if (/^[A-Za-z0-9_-]{{11}}$/.test(text)) return text;
      if (!text) return "";

      try {{
        const withProtocol = text.includes("://") ? text : `https://${{text}}`;
        const url = new URL(withProtocol);
        const host = url.hostname.replace(/^www\\./, "");
        const fromQuery = url.searchParams.get("v");
        if (fromQuery && /^[A-Za-z0-9_-]{{11}}$/.test(fromQuery)) return fromQuery;
        const parts = url.pathname.split("/").filter(Boolean);
        if (host === "youtu.be" && parts[0] && /^[A-Za-z0-9_-]{{11}}$/.test(parts[0])) return parts[0];
        for (const key of ["embed", "shorts", "live"]) {{
          const index = parts.indexOf(key);
          const id = index >= 0 ? parts[index + 1] : "";
          if (id && /^[A-Za-z0-9_-]{{11}}$/.test(id)) return id;
        }}
      }} catch (error) {{
        const match = text.match(/(?:v=|youtu\\.be\\/|embed\\/|shorts\\/|live\\/)([A-Za-z0-9_-]{{11}})/);
        if (match) return match[1];
      }}

      return "";
    }}

    function youtubeWatchUrl(videoId) {{
      return `https://www.youtube.com/watch?v=${{videoId}}`;
    }}

    function applyYoutubeUrl(rawValue, persist = true) {{
      const value = String(rawValue || "").trim();
      const videoId = extractYoutubeVideoId(value) || DEFAULT_YOUTUBE_VIDEO_ID;
      youtubeFrame.src = `https://www.youtube.com/embed/${{encodeURIComponent(videoId)}}`;
      youtubeUrlInput.value = value || youtubeWatchUrl(DEFAULT_YOUTUBE_VIDEO_ID);
      youtubeUrlStatus.textContent = `反映中: ${{youtubeWatchUrl(videoId)}}`;
      if (persist) {{
        if (value) {{
          localStorage.setItem(YOUTUBE_URL_STORAGE_KEY, value);
        }} else {{
          localStorage.removeItem(YOUTUBE_URL_STORAGE_KEY);
        }}
      }}
      return videoId;
    }}

    function applyYoutubeInput() {{
      const videoId = extractYoutubeVideoId(youtubeUrlInput.value);
      if (!videoId) {{
        youtubeUrlStatus.textContent = "YouTube URL または11文字の動画IDを入力してください。";
        youtubeUrlInput.focus();
        return;
      }}
      applyYoutubeUrl(youtubeUrlInput.value);
    }}

    function resetYoutubeUrl() {{
      localStorage.removeItem(YOUTUBE_URL_STORAGE_KEY);
      applyYoutubeUrl("", false);
      youtubeUrlStatus.textContent = `既定に戻しました: ${{youtubeWatchUrl(DEFAULT_YOUTUBE_VIDEO_ID)}}`;
    }}

    function applyLapColWidth(rawWidth, persist = true) {{
      const min = Number(lapColWidth.min) || 200;
      const max = Number(lapColWidth.max) || 680;
      const width = Math.min(max, Math.max(min, Number(rawWidth) || 260));
      lapColumn.style.setProperty("--lap-col-width", `${{width}}px`);
      lapColWidth.value = String(width);
      if (persist) localStorage.setItem(LAP_COL_WIDTH_STORAGE_KEY, String(width));
      clampLapColPos();
    }}

    function clampLapColPos() {{
      if (!lapColumn.style.left && !lapColumn.style.top) return;
      const rect = lapColumn.getBoundingClientRect();
      const maxLeft = Math.max(0, window.innerWidth - rect.width - 8);
      const maxTop = Math.max(0, window.innerHeight - 40);
      const left = Math.min(maxLeft, Math.max(0, parseFloat(lapColumn.style.left) || 0));
      const top = Math.min(maxTop, Math.max(0, parseFloat(lapColumn.style.top) || 0));
      lapColumn.style.left = `${{left}}px`;
      lapColumn.style.top = `${{top}}px`;
    }}

    function applyLapColPos(pos) {{
      if (pos && Number.isFinite(pos.left) && Number.isFinite(pos.top)) {{
        lapColumn.style.left = `${{pos.left}}px`;
        lapColumn.style.top = `${{pos.top}}px`;
        lapColumn.style.right = "auto";
      }} else {{
        lapColumn.style.left = "";
        lapColumn.style.top = "";
        lapColumn.style.right = "";
      }}
    }}

    function loadLapColPos() {{
      try {{
        const parsed = JSON.parse(localStorage.getItem(LAP_COL_POS_STORAGE_KEY) || "null");
        if (parsed && Number.isFinite(parsed.left) && Number.isFinite(parsed.top)) return parsed;
      }} catch (error) {{
        console.warn("周回カラム位置の復元に失敗", error);
      }}
      return null;
    }}

    function saveLapColPos() {{
      const left = parseFloat(lapColumn.style.left);
      const top = parseFloat(lapColumn.style.top);
      if (Number.isFinite(left) && Number.isFinite(top)) {{
        localStorage.setItem(LAP_COL_POS_STORAGE_KEY, JSON.stringify({{ left, top }}));
      }}
    }}

    function resetLapColPosition() {{
      localStorage.removeItem(LAP_COL_POS_STORAGE_KEY);
      localStorage.removeItem(LAP_COL_WIDTH_STORAGE_KEY);
      applyLapColPos(null);
      applyLapColWidth(260, false);
    }}

    function initLapColumnDrag() {{
      let dragging = null;
      lapColumnHandle.addEventListener("pointerdown", (event) => {{
        if (event.target.closest("input, button, .lap-filter-pills")) return;
        const rect = lapColumn.getBoundingClientRect();
        dragging = {{ dx: event.clientX - rect.left, dy: event.clientY - rect.top }};
        lapColumn.style.right = "auto";
        lapColumn.style.left = `${{rect.left}}px`;
        lapColumn.style.top = `${{rect.top}}px`;
        lapColumn.classList.add("dragging");
        lapColumnHandle.setPointerCapture(event.pointerId);
        event.preventDefault();
      }});
      lapColumnHandle.addEventListener("pointermove", (event) => {{
        if (!dragging) return;
        const rect = lapColumn.getBoundingClientRect();
        const maxLeft = Math.max(0, window.innerWidth - rect.width - 8);
        const maxTop = Math.max(0, window.innerHeight - 40);
        const left = Math.min(maxLeft, Math.max(0, event.clientX - dragging.dx));
        const top = Math.min(maxTop, Math.max(0, event.clientY - dragging.dy));
        lapColumn.style.left = `${{left}}px`;
        lapColumn.style.top = `${{top}}px`;
      }});
      const endDrag = (event) => {{
        if (!dragging) return;
        dragging = null;
        lapColumn.classList.remove("dragging");
        try {{ lapColumnHandle.releasePointerCapture(event.pointerId); }} catch (error) {{}}
        saveLapColPos();
      }};
      lapColumnHandle.addEventListener("pointerup", endDrag);
      lapColumnHandle.addEventListener("pointercancel", endDrag);
    }}

    let selectedDriverSet = loadSelectedDrivers();

    function selectedDrivers() {{
      return Array.from(driverCheckboxes.querySelectorAll("input:checked")).map((input) => input.value);
    }}

    function saveSelectedDrivers() {{
      const drivers = selectedDrivers();
      selectedDriverSet = new Set(drivers);
      if (drivers.length) {{
        localStorage.setItem(DRIVER_STORAGE_KEY, JSON.stringify(drivers));
      }} else {{
        localStorage.removeItem(DRIVER_STORAGE_KEY);
      }}
    }}

    function rebuildDriverFilter() {{
      const previously = new Set(selectedDrivers().length ? selectedDrivers() : selectedDriverSet);
      const names = new Set();
      for (const row of latestData.cars || []) {{
        if (!matchesClass(row.car_class)) continue;
        if (row.current_driver && row.current_driver !== "-") names.add(row.current_driver);
      }}
      for (const lap of latestData.lap_history || []) {{
        if (matchesClass(lap.car_class || "")) names.add(lap.driver_name);
      }}
      const sorted = Array.from(names).sort();
      driverCheckboxes.innerHTML = sorted.map((name) => {{
        const checked = previously.has(name) ? "checked" : "";
        return `<label><input type="checkbox" value="${{escapeHtml(name)}}" ${{checked}}>${{escapeHtml(name)}}</label>`;
      }}).join("");
      for (const input of driverCheckboxes.querySelectorAll("input")) {{
        input.addEventListener("change", () => {{
          saveSelectedDrivers();
          filterRows();
        }});
      }}
      saveSelectedDrivers();
    }}

    function toggleDriver(driverName) {{
      const name = String(driverName || "").trim();
      if (!name || name === "-") return;
      for (const input of driverCheckboxes.querySelectorAll("input")) {{
        if (input.value === name) {{
          input.checked = !input.checked;
          saveSelectedDrivers();
          filterRows();
          return;
        }}
      }}
    }}

    function updateMeta() {{
      metaBox.innerHTML = [
        `<span class="meta-item">レース: ${{escapeHtml(latestData.race_name || "")}} (${{escapeHtml(latestData.race_year || "")}})</span>`,
        `<span class="meta-item">クラス: ${{escapeHtml(selectedClass())}} / 全 ${{latestData.cars?.length || 0}} 台</span>`,
        `<span class="meta-item">更新: ${{escapeHtml(latestData.update_time_display || "")}}</span>`,
        `<span class="meta-item">保存: ${{escapeHtml(latestData.saved_at || "")}}</span>`,
      ].join("");
      const isLive = (latestData.status || "").includes("受信中");
      statusBadge.textContent = latestData.status || "-";
      statusBadge.className = "status " + (isLive ? "status-live" : "status-wait");
    }}

    function renderRows() {{
      const rows = (latestData.cars || []).filter((row) => matchesClass(row.car_class));
      const html = rows.map((row) => {{
        const searchText = [row.car_no, row.current_driver, row.all_drivers, row.team_name, row.car_name].join(" ");
        const teamClass = TEAM_CAR_NOS.includes(String(row.car_no)) ? "row-team" : "";
        const liveClass = row.has_live_data ? "row-live" : "";
        const rowClass = [teamClass, liveClass].filter(Boolean).join(" ");
        const bestClass = row.best_lap_ms > 0 ? "lap-best" : "";
        return `<tr class="${{rowClass}}" data-driver="${{escapeHtml(row.current_driver)}}" data-class="${{escapeHtml(row.car_class)}}" data-search="${{escapeHtml(searchText)}}">
  <td class="num">${{escapeHtml(row.pos)}}</td>
  <td class="num">${{escapeHtml(row.pic)}}</td>
  <td class="num car-no">${{escapeHtml(row.car_no)}}</td>
  <td><span class="class-badge">${{escapeHtml(row.car_class)}}</span></td>
  <td class="driver driver-pickable" title="ダブルクリックで選択切替">${{escapeHtml(row.current_driver)}}</td>
  <td class="drivers">${{escapeHtml(row.all_drivers)}}</td>
  <td class="lap ${{bestClass}}">${{escapeHtml(row.best_lap)}}</td>
  <td class="lap">${{escapeHtml(row.last_lap)}}</td>
  <td class="lap sector">${{escapeHtml(row.s1 || "-")}}</td>
  <td class="lap sector">${{escapeHtml(row.s2 || "-")}}</td>
  <td class="lap sector">${{escapeHtml(row.s3 || "-")}}</td>
  <td class="num">${{escapeHtml(row.laps)}}</td>
  <td class="team">${{escapeHtml(row.team_name)}}</td>
  <td class="car">${{escapeHtml(row.car_name)}}</td>
</tr>`;
      }}).join("");
      const scrollTop = tableWrap.scrollTop;
      const scrollLeft = tableWrap.scrollLeft;
      rowsBody.innerHTML = html;
      tableWrap.scrollTop = scrollTop;
      tableWrap.scrollLeft = scrollLeft;
      filterRows();
    }}

    function updateLapSortHeaders() {{
      for (const header of lapSortHeaders) {{
        const active = header.dataset.lapSort === lapSort.key;
        header.classList.toggle("sort-asc", active && lapSort.dir === 1);
        header.classList.toggle("sort-desc", active && lapSort.dir === -1);
      }}
    }}

    function lapSortValue(lap, key) {{
      if (key === "time") return Number(lap.lap_time_ms) || 0;
      if (key === "driver") return String(lap.driver_name || "");
      if (key === "slot") return Number(lap.driver_slot) || 0;
      return Number(lap.lap_no) || 0;
    }}

    function compareLapValues(a, b, key) {{
      const av = lapSortValue(a, key);
      const bv = lapSortValue(b, key);
      if (typeof av === "string" || typeof bv === "string") {{
        return String(av).localeCompare(String(bv), "ja");
      }}
      return av - bv;
    }}

    function visibleLapDrivers(drivers) {{
      const allowed = new Set(drivers);
      lapDriverFilterSet = new Set(Array.from(lapDriverFilterSet).filter((name) => allowed.has(name)));
      if (!lapDriverFilterSet.size) return drivers;
      const filtered = drivers.filter((name) => lapDriverFilterSet.has(name));
      return filtered.length ? filtered : drivers;
    }}

    function renderLapDriverFilter(drivers) {{
      if (drivers.length <= 1) {{
        lapFilterPills.innerHTML = "";
        return;
      }}
      const active = new Set(visibleLapDrivers(drivers));
      lapFilterPills.innerHTML = drivers.map((name) => {{
        const activeClass = active.has(name) ? "active" : "";
        return `<button type="button" class="${{activeClass}}" data-driver="${{escapeHtml(name)}}">${{escapeHtml(name)}}</button>`;
      }}).join("");
    }}

    function renderLapHistory() {{
      const drivers = selectedDrivers();
      if (!drivers.length) {{
        lapColumn.classList.add("hidden");
        return;
      }}

      const visibleDrivers = visibleLapDrivers(drivers);
      const driverSet = new Set(visibleDrivers);
      const sortLaps = (list) => list.slice().sort((a, b) => {{
        const primary = lapSort.dir * compareLapValues(a, b, lapSort.key);
        if (primary !== 0) return primary;
        const byLap = Number(a.lap_no) - Number(b.lap_no);
        return byLap !== 0 ? byLap : String(a.driver_name).localeCompare(String(b.driver_name), "ja");
      }});
      const laps = sortLaps(
        (latestData.lap_history || []).filter(
          (lap) => driverSet.has(lap.driver_name) && matchesClass(lap.car_class || "")
        )
      );

      lapColumn.classList.remove("hidden");
      lapPanelTitle.textContent = drivers.length === 1
        ? `${{drivers[0]}} の周回タイム`
        : `${{visibleDrivers.join(" / ")}} の周回タイム`;
      const multiDriver = drivers.length > 1;
      lapDriverCol.classList.toggle("hidden", !multiDriver);
      lapGroupBtn.classList.toggle("hidden", !multiDriver);
      lapGroupBtn.classList.toggle("active", multiDriver && lapGroupByDriver);
      renderLapDriverFilter(drivers);
      updateLapSortHeaders();
      if (!laps.length) {{
        lapPanelSummary.textContent = "まだ周回データがありません（取得開始後に蓄積されます）";
        lapRowsBody.innerHTML = "";
        return;
      }}

      const validMs = laps.map((lap) => Number(lap.lap_time_ms)).filter((v) => v > 0);
      const bestMs = Math.min(...validMs);
      const avgMs = validMs.length
        ? validMs.reduce((sum, v) => sum + v, 0) / validMs.length
        : 0;
      const avgDisplay = avgMs ? `${{(avgMs / 1000).toFixed(3)}}秒` : "-";
      lapPanelSummary.textContent = `全 ${{laps.length}} 周 / 平均 ${{avgDisplay}}`;
      const scrollTop = lapTableWrap.scrollTop;
      const showDriver = multiDriver;
      const colSpan = showDriver ? 4 : 3;
      const lapRowHtml = (lap) => {{
        const lapMs = Number(lap.lap_time_ms);
        const bestClass = lapMs === bestMs ? "lap-best" : "";
        const driverCell = showDriver
          ? `<td class="driver lap-driver-col">${{escapeHtml(lap.driver_name)}}</td>`
          : "";
        return `<tr>
  <td class="num">${{escapeHtml(lap.lap_no)}}</td>
  <td class="lap ${{bestClass}}">${{escapeHtml(lap.lap_time)}}</td>
  ${{driverCell}}
  <td class="num">${{escapeHtml(lap.driver_slot)}}</td>
</tr>`;
      }};

      if (showDriver && lapGroupByDriver) {{
        const sections = [];
        for (const name of visibleDrivers) {{
          const driverLaps = sortLaps(laps.filter((lap) => lap.driver_name === name));
          if (!driverLaps.length) continue;
          sections.push(`<tr class="lap-group-row"><td colspan="${{colSpan}}">${{escapeHtml(name)}}（${{driverLaps.length}}周）</td></tr>`);
          sections.push(driverLaps.map(lapRowHtml).join(""));
        }}
        lapRowsBody.innerHTML = sections.join("");
      }} else {{
        lapRowsBody.innerHTML = laps.map(lapRowHtml).join("");
      }}
      lapTableWrap.scrollTop = scrollTop;
    }}

    function filterRows() {{
      const drivers = selectedDrivers();
      const driverSet = new Set(drivers);
      const query = search.value.trim().toLowerCase();
      let visible = 0;
      for (const row of rowsBody.querySelectorAll("tr")) {{
        const driverName = row.dataset.driver || "";
        const haystack = (row.dataset.search || "").toLowerCase();
        const matchDriver = !driverSet.size || driverSet.has(driverName);
        const matchSearch = !query || haystack.includes(query);
        const show = matchDriver && matchSearch;
        row.classList.toggle("hidden", !show);
        if (show) visible++;
      }}
      count.textContent = String(visible);
      renderLapHistory();
      updateMeta();
    }}

    function applyData(data) {{
      latestData = data;
      rebuildClassFilter();
      rebuildDriverFilter();
      renderRows();
    }}

    async function refreshData() {{
      try {{
        const response = await fetch(`${{DATA_URL}}?t=${{Date.now()}}`, {{ cache: "no-store" }});
        if (!response.ok) return;
        applyData(await response.json());
      }} catch (error) {{
        console.warn("更新に失敗しました", error);
        if (location.protocol === "file:" && !(latestData.cars || []).length) {{
          statusBadge.textContent = "file:// では更新不可";
          statusBadge.className = "status status-wait";
          metaBox.innerHTML += `<span class="meta-item"><strong>start.bat を起動すると自動更新できます。</strong></span>`;
        }}
      }}
    }}

    classFilter.addEventListener("change", () => {{
      localStorage.setItem(CLASS_STORAGE_KEY, classFilter.value);
      rebuildDriverFilter();
      renderRows();
    }});
    driverClearBtn.addEventListener("click", () => {{
      for (const input of driverCheckboxes.querySelectorAll("input")) input.checked = false;
      saveSelectedDrivers();
      filterRows();
    }});
    search.addEventListener("input", filterRows);
    rowsBody.addEventListener("dblclick", (event) => {{
      const cell = event.target.closest("td.driver");
      if (!cell) return;
      const row = cell.closest("tr");
      toggleDriver(row?.dataset.driver || cell.textContent);
    }});
    lapFilterPills.addEventListener("click", (event) => {{
      const button = event.target.closest("button[data-driver]");
      if (!button) return;
      const drivers = selectedDrivers();
      if (!lapDriverFilterSet.size) lapDriverFilterSet = new Set(drivers);
      const name = button.dataset.driver;
      if (lapDriverFilterSet.has(name)) {{
        lapDriverFilterSet.delete(name);
      }} else {{
        lapDriverFilterSet.add(name);
      }}
      if (!lapDriverFilterSet.size) lapDriverFilterSet = new Set(drivers);
      saveLapDriverFilter();
      renderLapHistory();
    }});
    for (const header of lapSortHeaders) {{
      header.addEventListener("click", () => {{
        const key = header.dataset.lapSort;
        if (lapSort.key === key) {{
          lapSort.dir *= -1;
        }} else {{
          lapSort = {{ key, dir: 1 }};
        }}
        saveLapSort();
        renderLapHistory();
      }});
    }}
    videoSize.addEventListener("input", () => applyVideoSize(videoSize.value));
    applyVideoSize(localStorage.getItem(VIDEO_SIZE_STORAGE_KEY), false);
    youtubeApplyBtn.addEventListener("click", applyYoutubeInput);
    youtubeResetBtn.addEventListener("click", resetYoutubeUrl);
    youtubeUrlInput.addEventListener("keydown", (event) => {{
      if (event.key === "Enter") applyYoutubeInput();
    }});
    applyYoutubeUrl(localStorage.getItem(YOUTUBE_URL_STORAGE_KEY) || "", false);
    lapColWidth.addEventListener("input", () => applyLapColWidth(lapColWidth.value));
    lapColReset.addEventListener("click", resetLapColPosition);
    lapGroupBtn.addEventListener("click", () => {{
      lapGroupByDriver = !lapGroupByDriver;
      localStorage.setItem(LAP_GROUP_STORAGE_KEY, lapGroupByDriver ? "1" : "0");
      renderLapHistory();
    }});
    window.addEventListener("resize", clampLapColPos);
    initLapColumnDrag();
    applyLapColWidth(localStorage.getItem(LAP_COL_WIDTH_STORAGE_KEY), false);
    applyLapColPos(loadLapColPos());
    updateLapSortHeaders();
    if (window.__INITIAL_VIEW_DATA__) {{
      applyData(window.__INITIAL_VIEW_DATA__);
    }}
    refreshData();
    setInterval(refreshData, POLL_INTERVAL);
  </script>
</body>
</html>
"""

    html_path.write_text(page, encoding="utf-8")
    return html_path


def save_drivers_html(
    output_dir: Path,
    interval: int,
    default_class: str,
    initial_view_data: dict[str, Any] | None = None,
) -> Path:
    ensure_output_dir(output_dir)
    html_path = output_dir / "drivers.html"
    data_url = view_data_path(output_dir).name

    page = f"""<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{html.escape(TEAM_NAME)} | ドライバー一覧</title>
  <style>{dashboard_theme_css(drivers_page=True)}</style>
</head>
<body>
  <div class="wrap">
    {dashboard_header_html(f"{TEAM_NAME} ドライバー一覧")}

    <div class="toolbar">
      <select id="classFilter"></select>
      <select id="carNoFilter">
        <option value="">車番で絞り込み</option>
      </select>
      <select id="driverFilter">
        <option value="">走行中ドライバー一覧</option>
      </select>
      <input id="search" type="search" placeholder="ドライバー・車番・チーム名で検索...">
      <a class="nav-link" href="index.html">← ダッシュボードに戻る</a>
      <a class="nav-link" href="timing_live.xlsx" download>Excel出力</a>
      <div class="info">走行中: <span id="count">0</span> 件 / <span id="pollInfo">{interval}秒ごとに自動更新</span></div>
    </div>

    <div class="table-wrap" id="tableWrap">
      <table>
        <thead>
          <tr>
            <th class="sticky-col sticky-col-1">Driver</th>
            <th class="sticky-col sticky-col-2">Slot</th>
            <th class="sticky-col sticky-col-3">状態</th>
            <th>No.</th>
            <th>Class</th>
            <th>BestLap</th>
            <th>LastLap</th>
            <th>S1</th>
            <th>S2</th>
            <th>S3</th>
            <th>Laps</th>
            <th>POS</th>
            <th>PIC</th>
            <th>Team</th>
            <th>Car</th>
          </tr>
        </thead>
        <tbody id="rows"></tbody>
      </table>
    </div>

    <div id="lapPanel" class="lap-panel hidden">
      <h2 id="lapPanelTitle">周回タイム</h2>
      <div class="summary" id="lapPanelSummary"></div>
      <div class="table-wrap" id="lapTableWrap">
        <table>
          <thead>
            <tr>
              <th id="lapSortHeader" class="sortable sort-asc">Lap<span class="sort-icon"></span></th>
              <th>Time</th>
              <th id="driverTh">Driver</th>
              <th>No.</th>
              <th>Slot</th>
              <th>記録時刻</th>
            </tr>
          </thead>
          <tbody id="lapRows"></tbody>
        </table>
      </div>
    </div>

    <div class="footer">
      {html.escape(TEAM_NAME)} 専用ダッシュボード / ドライバー選択で周回タイムを表示
    </div>
  </div>

  {embed_view_data_script(initial_view_data)}
  <script>
    const DATA_URL = "{html.escape(data_url)}";
    const DEFAULT_CLASS = "{html.escape(default_class)}";
    const TEAM_CAR_NOS = {json.dumps(list(TEAM_CAR_NOS))};
    const CLASS_STORAGE_KEY = "st_selected_class";
    const POLL_INTERVAL = {interval} * 1000;
    const classFilter = document.getElementById("classFilter");
    const carNoFilter = document.getElementById("carNoFilter");
    const driverFilter = document.getElementById("driverFilter");
    const driverTh = document.getElementById("driverTh");
    const search = document.getElementById("search");
    const rowsBody = document.getElementById("rows");
    const count = document.getElementById("count");
    const tableWrap = document.getElementById("tableWrap");
    const lapPanel = document.getElementById("lapPanel");
    const lapPanelTitle = document.getElementById("lapPanelTitle");
    const lapPanelSummary = document.getElementById("lapPanelSummary");
    const lapRowsBody = document.getElementById("lapRows");
    const lapTableWrap = document.getElementById("lapTableWrap");
    const lapSortHeader = document.getElementById("lapSortHeader");
    const metaBox = document.getElementById("metaBox");
    const statusBadge = document.getElementById("statusBadge");
    let latestData = {{ drivers: [], lap_history: [] }};
    let lapSortDir = 1;

    function escapeHtml(value) {{
      return String(value)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");
    }}

    function selectedClass() {{
      return classFilter.value || "ALL";
    }}

    function matchesClass(carClass) {{
      const selected = selectedClass();
      return selected === "ALL" || carClass === selected;
    }}

    function runningDrivers() {{
      return (latestData.drivers || []).filter((row) => row.is_current && matchesClass(row.car_class));
    }}

    function rebuildClassFilter() {{
      const selected = classFilter.value || localStorage.getItem(CLASS_STORAGE_KEY) || latestData.default_class || DEFAULT_CLASS;
      classFilter.innerHTML = '<option value="ALL">全クラス</option>';
      for (const carClass of latestData.classes || []) {{
        const option = document.createElement("option");
        option.value = carClass;
        option.textContent = carClass;
        classFilter.appendChild(option);
      }}
      const values = Array.from(classFilter.options).map((option) => option.value);
      classFilter.value = values.includes(selected) ? selected : "ALL";
    }}

    function updateMeta() {{
      metaBox.innerHTML = [
        `<span class="meta-item">レース: ${{escapeHtml(latestData.race_name || "")}} (${{escapeHtml(latestData.race_year || "")}})</span>`,
        `<span class="meta-item">クラス: ${{escapeHtml(selectedClass())}}</span>`,
        `<span class="meta-item">更新: ${{escapeHtml(latestData.update_time_display || "")}}</span>`,
        `<span class="meta-item">保存: ${{escapeHtml(latestData.saved_at || "")}}</span>`,
      ].join("");
      const isLive = (latestData.status || "").includes("受信中");
      statusBadge.textContent = latestData.status || "-";
      statusBadge.className = "status " + (isLive ? "status-live" : "status-wait");
    }}

    function rebuildDriverFilter() {{
      const selected = driverFilter.value;
      const names = new Set();
      for (const row of runningDrivers()) names.add(row.driver_name);
      for (const lap of latestData.lap_history || []) {{
        if (matchesClass(lap.car_class || "")) names.add(lap.driver_name);
      }}
      const sorted = Array.from(names).sort();
      driverFilter.innerHTML = '<option value="">走行中ドライバー一覧</option>';
      for (const name of sorted) {{
        const option = document.createElement("option");
        option.value = name;
        option.textContent = name;
        driverFilter.appendChild(option);
      }}
      if (selected && sorted.includes(selected)) driverFilter.value = selected;
    }}

    function rebuildCarNoFilter() {{
      const selected = carNoFilter.value;
      const carNos = new Set();
      // 走行中ドライバーの車番を収集
      for (const row of latestData.drivers || []) {{
        if (matchesClass(row.car_class)) carNos.add(String(row.car_no));
      }}
      // ラップ履歴の車番を収集
      for (const lap of latestData.lap_history || []) {{
        if (matchesClass(lap.car_class || "")) carNos.add(String(lap.car_no));
      }}
      const sorted = Array.from(carNos).sort((a, b) => Number(a) - Number(b));
      carNoFilter.innerHTML = '<option value="">車番で絞り込み</option>';
      for (const no of sorted) {{
        const option = document.createElement("option");
        option.value = no;
        option.textContent = `#${{no}}`;
        carNoFilter.appendChild(option);
      }}
      if (selected && sorted.includes(selected)) carNoFilter.value = selected;
    }}

    function renderRunningTable() {{
      const rows = runningDrivers();
      const html = rows.map((row) => {{
        const searchText = [row.driver_name, row.car_no, row.team_name, row.car_name].join(" ");
        const bestClass = row.best_lap_ms > 0 ? "lap-best" : "";
        const teamClass = TEAM_CAR_NOS.includes(String(row.car_no)) ? "row-team" : "";
        const rowClass = ["row-live", teamClass].filter(Boolean).join(" ");
        return `<tr class="${{rowClass}}" data-driver="${{escapeHtml(row.driver_name)}}" data-carno="${{escapeHtml(String(row.car_no))}}" data-class="${{escapeHtml(row.car_class)}}" data-search="${{escapeHtml(searchText)}}">
  <td class="driver sticky-col sticky-col-1">${{escapeHtml(row.driver_name)}}</td>
  <td class="num sticky-col sticky-col-2">${{escapeHtml(row.driver_slot)}}</td>
  <td class="current sticky-col sticky-col-3"><span class="current-badge">走行中</span></td>
  <td class="num car-no">${{escapeHtml(row.car_no)}}</td>
  <td><span class="class-badge">${{escapeHtml(row.car_class)}}</span></td>
  <td class="lap ${{bestClass}}">${{escapeHtml(row.best_lap)}}</td>
  <td class="lap">${{escapeHtml(row.last_lap)}}</td>
  <td class="lap sector">${{escapeHtml(row.s1 || "-")}}</td>
  <td class="lap sector">${{escapeHtml(row.s2 || "-")}}</td>
  <td class="lap sector">${{escapeHtml(row.s3 || "-")}}</td>
  <td class="num">${{escapeHtml(row.laps)}}</td>
  <td class="num">${{escapeHtml(row.pos)}}</td>
  <td class="num">${{escapeHtml(row.pic)}}</td>
  <td class="team">${{escapeHtml(row.team_name)}}</td>
  <td class="car">${{escapeHtml(row.car_name)}}</td>
</tr>`;
      }}).join("");
      const scrollTop = tableWrap.scrollTop;
      const scrollLeft = tableWrap.scrollLeft;
      rowsBody.innerHTML = html;
      tableWrap.scrollTop = scrollTop;
      tableWrap.scrollLeft = scrollLeft;
      filterRows();
    }}

    function updateLapSortHeader() {{
      lapSortHeader.classList.toggle("sort-asc", lapSortDir === 1);
      lapSortHeader.classList.toggle("sort-desc", lapSortDir === -1);
    }}

    function renderLapHistory() {{
      const selectedDriver = driverFilter.value;
      const selectedCarNo = carNoFilter.value;

      // 車番・ドライバーどちらも未選択なら非表示
      if (!selectedDriver && !selectedCarNo) {{
        lapPanel.classList.add("hidden");
        return;
      }}

      // 車番フィルタ時はドライバー列を表示、ドライバーフィルタのみの場合は非表示
      const showDriverCol = !!selectedCarNo;
      driverTh.classList.toggle("hidden", !showDriverCol);

      const laps = (latestData.lap_history || [])
        .filter((lap) => {{
          const classOk = matchesClass(lap.car_class || "");
          const driverOk = !selectedDriver || lap.driver_name === selectedDriver;
          const carOk = !selectedCarNo || String(lap.car_no) === selectedCarNo;
          return classOk && driverOk && carOk;
        }})
        .sort((a, b) => lapSortDir * (Number(a.lap_no) - Number(b.lap_no)));

      lapPanel.classList.remove("hidden");
      const label = selectedCarNo ? `#${{selectedCarNo}}` : selectedDriver;
      lapPanelTitle.textContent = `${{label}} の周回タイム`;
      if (!laps.length) {{
        lapPanelSummary.textContent = "まだ周回データがありません（取得開始後に蓄積されます）";
        lapRowsBody.innerHTML = "";
        return;
      }}

      const bestMs = Math.min(...laps.map((lap) => Number(lap.lap_time_ms)).filter((v) => v > 0));
      lapPanelSummary.textContent = `全 ${{laps.length}} 周`;
      const scrollTop = lapTableWrap.scrollTop;
      lapRowsBody.innerHTML = laps.map((lap) => {{
        const bestClass = Number(lap.lap_time_ms) === bestMs ? "lap-best" : "";
        const driverCell = showDriverCol
          ? `<td class="driver">${{escapeHtml(lap.driver_name || "")}}</td>`
          : `<td class="hidden"></td>`;
        return `<tr>
  <td class="num">${{escapeHtml(lap.lap_no)}}</td>
  <td class="lap ${{bestClass}}">${{escapeHtml(lap.lap_time)}}</td>
  ${{driverCell}}
  <td class="num car-no">${{escapeHtml(lap.car_no)}}</td>
  <td class="num">${{escapeHtml(lap.driver_slot)}}</td>
  <td>${{escapeHtml(lap.recorded_at || "")}}</td>
</tr>`;
      }}).join("");
      lapTableWrap.scrollTop = scrollTop;
    }}

    function filterRows() {{
      const selectedDriver = driverFilter.value;
      const selectedCarNo = carNoFilter.value;
      const query = search.value.trim().toLowerCase();
      let visible = 0;
      for (const row of rowsBody.querySelectorAll("tr")) {{
        const driverName = row.dataset.driver || "";
        const rowCarNo = row.dataset.carno || "";
        const haystack = (row.dataset.search || "").toLowerCase();
        const matchDriver = !selectedDriver || driverName === selectedDriver;
        const matchCarNo = !selectedCarNo || rowCarNo === selectedCarNo;
        const matchSearch = !query || haystack.includes(query);
        const show = matchDriver && matchCarNo && matchSearch;
        row.classList.toggle("hidden", !show);
        if (show) visible++;
      }}
      count.textContent = String(visible);
      renderLapHistory();
      updateMeta();
    }}

    function applyData(data) {{
      latestData = data;
      rebuildClassFilter();
      rebuildDriverFilter();
      rebuildCarNoFilter();
      renderRunningTable();
      updateMeta();
    }}

    async function refreshData() {{
      try {{
        const response = await fetch(`${{DATA_URL}}?t=${{Date.now()}}`, {{ cache: "no-store" }});
        if (!response.ok) return;
        applyData(await response.json());
      }} catch (error) {{
        console.warn("更新に失敗しました", error);
      }}
    }}

    classFilter.addEventListener("change", () => {{
      localStorage.setItem(CLASS_STORAGE_KEY, classFilter.value);
      rebuildDriverFilter();
      rebuildCarNoFilter();
      renderRunningTable();
      updateMeta();
    }});
    // 車番を選んだらドライバーフィルタをリセット（排他）
    carNoFilter.addEventListener("change", () => {{
      if (carNoFilter.value) driverFilter.value = "";
      filterRows();
    }});
    // ドライバーを選んだら車番フィルタをリセット（排他）
    driverFilter.addEventListener("change", () => {{
      if (driverFilter.value) carNoFilter.value = "";
      filterRows();
    }});
    search.addEventListener("input", filterRows);
    lapSortHeader.addEventListener("click", () => {{
      lapSortDir *= -1;
      updateLapSortHeader();
      renderLapHistory();
    }});
    updateLapSortHeader();
    refreshData();
    setInterval(refreshData, POLL_INTERVAL);
  </script>
</body>
</html>
"""

    html_path.write_text(page, encoding="utf-8")
    return html_path


def persist_outputs(
    args: argparse.Namespace,
    master: dict[str, Any],
    live: dict[str, Any],
) -> dict[str, Path]:
    update_sector_state(args.output_dir, live)
    sector_state = read_json_file(sector_state_path(args.output_dir))
    all_rows = build_rows(master, live, None, args.english, sector_state=sector_state)
    all_driver_rows = build_driver_rows(
        master, live, None, args.english, sector_state=sector_state
    )
    default_class = class_label_for(args.class_filter)

    saved_paths = save_latest_files(args.output_dir, all_rows, master, live)
    if args.history:
        saved_paths["history"] = append_history(args.output_dir, all_rows, master, live)
        saved_paths["raw_snapshots"] = append_raw_snapshot(args.output_dir, master, live)
    if args.excel:
        saved_paths["drivers_running"] = save_drivers_running_csv(
            args.output_dir,
            all_driver_rows,
            master,
            live,
        )
        excel_path = export_live_xlsx(
            args.output_dir,
            all_rows,
            all_driver_rows,
            master,
            live,
            default_class,
            args.interval,
        )
        if excel_path:
            saved_paths["excel"] = excel_path
    if args.html:
        saved_paths["lap_history"] = update_lap_history(
            args.output_dir,
            master,
            live,
            args.english,
        )
        view_data_path = save_view_data_json(
            args.output_dir,
            all_rows,
            all_driver_rows,
            master,
            live,
            args.interval,
            default_class,
        )
        saved_paths["view_data"] = view_data_path
        view_data = read_json_file(view_data_path)
        saved_paths["html"] = save_html_file(
            args.output_dir,
            args.interval,
            default_class,
            initial_view_data=view_data,
        )
        saved_paths["drivers_html"] = save_drivers_html(
            args.output_dir,
            args.interval,
            default_class,
            initial_view_data=view_data,
        )
    return saved_paths


def run_once(args: argparse.Namespace) -> int:
    session = requests.Session()
    session.headers.update({"User-Agent": "supertaikyu-timing-fetcher/1.0"})

    try:
        rows, master, live = load_timing_data(session, None, args.english)
    except RuntimeError as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1

    display_rows = filter_rows_by_class(rows, args.class_filter)
    if args.search:
        print_search_results(search_rows(display_rows, args.search), args.search)
        return 0

    saved_paths = persist_outputs(args, master, live)

    if args.html:
        print(f"HTML: {saved_paths['html']}")
    if not args.quiet:
        header = [
            "スーパー耐久 ライブタイミング取得ツール",
            f"状態: {race_status_text(live)}",
            f"更新時刻: {format_update_time(live.get('UpdateTime', ''))}",
            f"レース: {master.get('RaceNameL', '-')} ({master.get('RaceYear', '-')})",
            f"表示クラス: {class_label_for(args.class_filter)}",
            f"保存: 全クラス一括",
            f"保存先: {args.output_dir}",
            f"JSON: {saved_paths['json']}",
            f"CSV : {saved_paths['csv']}",
        ]
        if "view_data" in saved_paths:
            header.append(f"View JSON: {saved_paths['view_data']}")
        if "html" in saved_paths:
            header.append(f"HTML: {saved_paths['html']}")
        if "drivers_html" in saved_paths:
            header.append(f"Drivers HTML: {saved_paths['drivers_html']}")
        if "excel" in saved_paths:
            header.append(f"Excel: {saved_paths['excel']}")
        if "raw_snapshots" in saved_paths:
            header.append(f"Raw CSV: {saved_paths['raw_snapshots']}")
        print_table(display_rows, header)
    return 0


def run_serve_browser(args: argparse.Namespace) -> int:
    page = "data/drivers.html" if args.serve_page == "drivers" else "data/index.html"
    server = start_static_server(APP_DIR, args.http_port)
    url = dashboard_page_url(args.http_port, page)
    webbrowser.open(url)
    print(f"ローカルサーバー: http://127.0.0.1:{args.http_port}/")
    print(f"ブラウザを開きました: {url}")
    print("終了は Ctrl+C です。")
    try:
        while True:
            time.sleep(3600)
    except KeyboardInterrupt:
        print("\n終了します。")
        server.shutdown()
        return 0


def run_watch(args: argparse.Namespace) -> int:
    session = requests.Session()
    session.headers.update({"User-Agent": "supertaikyu-timing-fetcher/1.0"})

    browser_opened = False
    if args.html:
        start_static_server(APP_DIR, args.http_port)
        if not args.quiet:
            print(f"ダッシュボード: {dashboard_page_url(args.http_port)}")

    if args.html and not args.quiet:
        print("HTML 表示モードを開始します。ブラウザを見ながらバックグラウンド更新します。")
        print("終了は Ctrl+C です。")
        time.sleep(1)
    elif not args.quiet:
        print("自動更新モードを開始します。終了は Ctrl+C です。")
        time.sleep(1)

    while True:
        try:
            rows, master, live = load_timing_data(session, None, args.english)
            display_rows = filter_rows_by_class(rows, args.class_filter)
            if args.search:
                display_rows = search_rows(display_rows, args.search)

            saved_paths = persist_outputs(args, master, live)

            if args.open_browser and args.html and not browser_opened and "html" in saved_paths:
                url = dashboard_page_url(args.http_port)
                webbrowser.open(url)
                browser_opened = True
                if not args.quiet:
                    print(f"ブラウザで HTML を開きました: {url}")

            if args.quiet or args.html:
                status = race_status_text(live)
                updated = format_update_time(live.get("UpdateTime", ""))
                html_info = ""
                if "html" in saved_paths:
                    html_info = f" | HTML: {saved_paths['html']}"
                if "drivers_html" in saved_paths:
                    html_info += f" | Drivers: {saved_paths['drivers_html']}"
                if "excel" in saved_paths:
                    html_info += f" | Excel: {saved_paths['excel']}"
                print(
                    f"[{datetime.now().strftime('%H:%M:%S')}] {status} | 更新: {updated} | 全{len(rows)}台 / 表示{len(display_rows)}台{html_info}"
                )
            else:
                clear_screen()
                header = [
                    "スーパー耐久 ライブタイミング取得ツール",
                    f"状態: {race_status_text(live)}",
                    f"更新時刻: {format_update_time(live.get('UpdateTime', ''))}",
                    f"レース: {master.get('RaceNameL', '-')} ({master.get('RaceYear', '-')})",
                    f"表示クラス: {class_label_for(args.class_filter)}",
                    f"保存: 全クラス一括 ({len(rows)}台)",
                    f"自動更新: {args.interval}秒",
                ]
                if args.search:
                    header.append(f"検索: {args.search}")
                if "history" in saved_paths:
                    header.append(f"履歴: {saved_paths['history']}")
                header.append(f"保存先: {args.output_dir}")
                header.append(f"JSON: {saved_paths['json']}")
                header.append(f"CSV : {saved_paths['csv']}")
                if "view_data" in saved_paths:
                    header.append(f"View JSON: {saved_paths['view_data']}")
                if "html" in saved_paths:
                    header.append(f"HTML: {saved_paths['html']}")
                print_table(display_rows, header)
                print("Ctrl+C で終了")

            time.sleep(args.interval)
        except KeyboardInterrupt:
            print("\n終了します。")
            return 0
        except RuntimeError as exc:
            if not args.quiet and not args.html:
                clear_screen()
            print(f"取得エラー: {exc}")
            print(f"{args.interval}秒後に再試行します...")
            time.sleep(args.interval)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Fetch Super Taikyu live timing data from supertaikyu.live JSON.",
    )
    parser.add_argument(
        "--class",
        dest="class_filter",
        default=DEFAULT_CLASS,
        help=f"Default display class for terminal/HTML (default: {DEFAULT_CLASS}). Data is always saved for all classes. Use ALL to show every class.",
    )
    parser.add_argument(
        "--search",
        help="Search by driver, team, or car name (partial match).",
    )
    parser.add_argument(
        "--english",
        action="store_true",
        help="Use English driver names.",
    )
    parser.add_argument(
        "--interval",
        type=int,
        default=DEFAULT_INTERVAL,
        help=f"Auto-refresh interval in seconds (default: {DEFAULT_INTERVAL}).",
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="Fetch once and exit.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory to save JSON/CSV data (default: {DEFAULT_OUTPUT_DIR}).",
    )
    parser.add_argument(
        "--history",
        action="store_true",
        help="Append each fetch to history CSV and raw JSON warehouse CSV.",
    )
    parser.add_argument(
        "--excel",
        action="store_true",
        help="Update data/timing_live.xlsx and warehouse/drivers_running.csv.",
    )
    parser.add_argument(
        "--html",
        action="store_true",
        help="Generate data/index.html for browser viewing.",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Minimize terminal output.",
    )
    parser.add_argument(
        "--open-browser",
        action="store_true",
        help="Open index.html in the default browser on start.",
    )
    parser.add_argument(
        "--http-port",
        type=int,
        default=DEFAULT_HTTP_PORT,
        help=f"Local HTTP port for dashboard (default: {DEFAULT_HTTP_PORT}).",
    )
    parser.add_argument(
        "--serve-browser",
        action="store_true",
        help="Serve saved HTML/JSON over HTTP and open the browser (no fetch loop).",
    )
    parser.add_argument(
        "--serve-page",
        choices=("index", "drivers"),
        default="index",
        help="Page to open with --serve-browser (default: index).",
    )
    return parser


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")

    parser = build_parser()
    args = parser.parse_args()

    if str(args.class_filter).upper() == "ALL":
        args.class_filter = None
    elif args.class_filter not in ALL_CLASSES:
        print(
            f"警告: 未知のクラス '{args.class_filter}' です。データがない場合があります。",
            file=sys.stderr,
        )

    if args.serve_browser:
        return run_serve_browser(args)
    if args.once:
        return run_once(args)
    return run_watch(args)


if __name__ == "__main__":
    raise SystemExit(main())
